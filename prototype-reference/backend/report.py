"""
report.py — generate an Excel report for a given date.

Columns (per spec):
  Camera Name | Date | Time | Identified Person | Reference Photo | Image

Image column holds an **embedded thumbnail** of the best face from the event
(not a hyperlink — actual JPEG pasted into the cell). The Reference Photo
column holds a clickable hyperlink to the first photo in the person's
known_people/ folder.

One row per event. Unknown events still show up (person column empty) so
reviewers can see that someone appeared but couldn't be identified.
"""

import argparse
import io
import sqlite3
import sys
import time
from pathlib import Path
from typing import Optional

import cv2
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from db import DB_PATH, init_db
from known_people import KNOWN_DIR, IMAGE_EXTS


REPORTS_DIR = Path("reports")
REPORTS_DIR.mkdir(exist_ok=True)

# Visual tuning for the image column
THUMB_PX = 60            # square thumbnail size in the Excel cell
IMAGE_COL_WIDTH = 12     # Excel column-width units (~7px each)
ROW_HEIGHT_PT = 50       # points (~1.33 px each); fits THUMB_PX with margin


def _ref_photo_path(person_name: str) -> Optional[Path]:
    """Pick the first reference photo for a person, or None."""
    pdir = KNOWN_DIR / person_name
    if not pdir.exists():
        return None
    for f in sorted(pdir.iterdir()):
        if f.suffix.lower() in IMAGE_EXTS:
            return f
    return None


def _thumbnail_bytes(image_path: Path, size: int = THUMB_PX) -> Optional[bytes]:
    """Read a JPEG from disk, resize to a square thumbnail, return JPEG bytes."""
    if not image_path.exists():
        return None
    img = cv2.imread(str(image_path))
    if img is None:
        return None
    h, w = img.shape[:2]
    # Center-crop to square then resize
    side = min(h, w)
    y0 = (h - side) // 2
    x0 = (w - side) // 2
    square = img[y0:y0 + side, x0:x0 + side]
    thumb = cv2.resize(square, (size, size), interpolation=cv2.INTER_AREA)
    ok, buf = cv2.imencode(".jpg", thumb, [cv2.IMWRITE_JPEG_QUALITY, 85])
    return buf.tobytes() if ok else None


def generate_report(
    date: str,
    output: Optional[str] = None,
    include_unknown: bool = True,
) -> dict:
    """
    Build the Excel report for a single date.

    Args:
      date: YYYY-MM-DD
      output: explicit output path; if omitted we pick a timestamped name
      include_unknown: keep rows for unidentified events too (recommended)

    Returns a summary dict.
    """
    init_db()

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        where = ["date = ?"]
        params = [date]
        if not include_unknown:
            where.append("person_name IS NOT NULL")
        sql = f"""
            SELECT id, camera_name, date, started_at, person_name, match_score
            FROM events
            WHERE {' AND '.join(where)}
            ORDER BY started_at
        """
        events = conn.execute(sql, params).fetchall()
        # Collect best-face paths per event in one pass
        best_face_paths = {}
        for ev in events:
            row = conn.execute(
                """SELECT file_path FROM faces
                   WHERE event_id = ? ORDER BY quality DESC LIMIT 1""",
                (ev["id"],),
            ).fetchone()
            best_face_paths[ev["id"]] = row["file_path"] if row else None

    if not events:
        return {"ok": False, "error": f"no events found for {date}"}

    # --- Build the workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = f"{date} report"

    # Header
    headers = ["Camera Name", "Date", "Time",
               "Identified Person", "Reference Photo", "Image"]
    ws.append(headers)

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="2A4E7A")
    header_align = Alignment(horizontal="center", vertical="center")
    body_font = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0066CC", underline="single")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    # Column widths
    widths = {1: 22, 2: 12, 3: 12, 4: 22, 5: 26, 6: IMAGE_COL_WIDTH}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # --- Data rows ---
    matched_count = 0
    unknown_count = 0

    for row_idx, ev in enumerate(events, start=2):
        time_str = (ev["started_at"] or "")[11:19]  # HH:MM:SS

        ws.cell(row=row_idx, column=1, value=ev["camera_name"]).font = body_font
        ws.cell(row=row_idx, column=2, value=ev["date"]).font = body_font
        ws.cell(row=row_idx, column=3, value=time_str).font = body_font
        ws.cell(row=row_idx, column=1).alignment = left
        ws.cell(row=row_idx, column=2).alignment = center
        ws.cell(row=row_idx, column=3).alignment = center

        person = ev["person_name"]
        if person:
            matched_count += 1
            pc = ws.cell(row=row_idx, column=4, value=person)
            pc.font = body_font
            pc.alignment = left

            # Reference photo hyperlink
            ref_path = _ref_photo_path(person)
            if ref_path:
                rc = ws.cell(row=row_idx, column=5, value=ref_path.name)
                rc.hyperlink = ref_path.resolve().as_uri()
                rc.font = link_font
                rc.alignment = left
            else:
                ws.cell(row=row_idx, column=5, value="(no reference)").font = body_font
        else:
            unknown_count += 1
            pc = ws.cell(row=row_idx, column=4, value="(unknown)")
            pc.font = Font(name="Arial", size=10, italic=True, color="777777")
            pc.alignment = left

        # Row height before embedding image
        ws.row_dimensions[row_idx].height = ROW_HEIGHT_PT

        # Embed thumbnail
        face_path = best_face_paths.get(ev["id"])
        if face_path:
            thumb_bytes = _thumbnail_bytes(Path(face_path))
            if thumb_bytes:
                xl_img = XLImage(io.BytesIO(thumb_bytes))
                # Resize in points (openpyxl uses EMU; explicit px via width/height work)
                xl_img.width = THUMB_PX
                xl_img.height = THUMB_PX
                anchor_cell = f"{get_column_letter(6)}{row_idx}"
                ws.add_image(xl_img, anchor_cell)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # --- Output path ---
    if output:
        output_path = Path(output)
    else:
        ts = time.strftime("%H%M%S")
        output_path = REPORTS_DIR / f"report_{date}__{ts}.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "ok": True,
        "output_path": str(output_path.resolve()),
        "output_name": output_path.name,
        "rows": len(events),
        "matched": matched_count,
        "unknown": unknown_count,
        "date": date,
    }


def list_reports() -> list[dict]:
    """List previously generated Excel reports, newest first."""
    if not REPORTS_DIR.exists():
        return []
    files = [f for f in REPORTS_DIR.iterdir()
             if f.is_file() and f.suffix.lower() == ".xlsx"]
    files.sort(key=lambda f: f.stat().st_mtime, reverse=True)
    out = []
    for f in files:
        # Infer report kind from filename prefix (see generate_* below)
        kind = "event"
        if f.name.startswith("attendance_"):
            kind = "attendance"
        elif f.name.startswith("report_"):
            kind = "event"
        out.append({
            "name": f.name,
            "kind": kind,
            "size_kb": round(f.stat().st_size / 1024, 1),
            "mtime": f.stat().st_mtime,
        })
    return out


# =====================================================================
# Attendance report (Round 3)
# =====================================================================
# Shape: ONE row per identified person per day, with
#   Camera Name (in) | Employee Name | Employee ID | In Time | Out Time |
#   Date | Total Duration | Reference Photo | In Snapshot
#
# Strategy: earliest event on that date = in, latest event = out.
# If someone only appears on one camera the whole day, in_camera == out_camera.
# Unknowns are excluded by default — attendance only tracks identified people.
#
# Employee ID: we don't have one in the DB. To keep the format flexible without
# adding more schema, we derive it from the person's name (hashed) unless you
# later add an employees table. If you want real employee IDs we can extend
# known_people.py to store metadata per person.

def _employee_id_for(name: str) -> str:
    """Deterministic short ID from the name. Replace with real IDs if needed."""
    import hashlib
    h = hashlib.sha1(name.encode("utf-8")).hexdigest()
    return f"E{int(h[:6], 16) % 100000:05d}"


def generate_attendance_report(
    date: str,
    output: Optional[str] = None,
) -> dict:
    """Build the daily attendance Excel.

    One row per identified person. For each person we find:
      - earliest event (in_time, in_camera, in face thumbnail)
      - latest event (out_time, out_camera)
    """
    init_db()

    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        events = conn.execute(
            """SELECT id, camera_name, person_name, started_at, ended_at
               FROM events
               WHERE date = ? AND person_name IS NOT NULL
               ORDER BY person_name, started_at""",
            (date,),
        ).fetchall()

        # Pre-fetch best face for each event so we can embed the "in" thumbnail
        best_faces = {}
        for ev in events:
            row = conn.execute(
                """SELECT file_path FROM faces
                   WHERE event_id = ? ORDER BY quality DESC LIMIT 1""",
                (ev["id"],),
            ).fetchone()
            best_faces[ev["id"]] = row["file_path"] if row else None

    if not events:
        return {"ok": False, "error": f"no identified events for {date}"}

    # Aggregate per person
    per_person = {}  # name -> dict(events=[...], first, last)
    for ev in events:
        name = ev["person_name"]
        per_person.setdefault(name, []).append(ev)

    # --- Workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = f"{date} attendance"

    headers = ["Camera (In)", "Employee Name", "Employee ID",
               "In Time", "Out Time", "Date",
               "Total Duration", "Reference Photo", "In Snapshot"]
    ws.append(headers)

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="2A4E7A")
    header_align = Alignment(horizontal="center", vertical="center")
    body_font = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0066CC", underline="single")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font; c.fill = header_fill; c.alignment = header_align

    # Column widths (Excel units)
    widths = {1: 20, 2: 22, 3: 14, 4: 12, 5: 12, 6: 12,
              7: 14, 8: 26, 9: IMAGE_COL_WIDTH}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    row_idx = 2
    for name in sorted(per_person.keys()):
        evs = per_person[name]
        first = evs[0]    # earliest (events were ordered by started_at)
        last = evs[-1]

        def time_only(iso):
            return (iso or "")[11:19]

        in_time = time_only(first["started_at"])
        out_time = time_only(last["ended_at"] or last["started_at"])

        # Total duration = last.ended (or started) - first.started, as HH:MM:SS
        try:
            from datetime import datetime as _dt
            t_in = _dt.fromisoformat(first["started_at"])
            t_out = _dt.fromisoformat(last["ended_at"] or last["started_at"])
            delta = t_out - t_in
            total = str(delta).split(".")[0]
        except Exception:
            total = ""

        ws.cell(row=row_idx, column=1, value=first["camera_name"]).font = body_font
        ws.cell(row=row_idx, column=2, value=name).font = body_font
        ws.cell(row=row_idx, column=3, value=_employee_id_for(name)).font = body_font
        ws.cell(row=row_idx, column=4, value=in_time).font = body_font
        ws.cell(row=row_idx, column=5, value=out_time).font = body_font
        ws.cell(row=row_idx, column=6, value=date).font = body_font
        ws.cell(row=row_idx, column=7, value=total).font = body_font
        for col_idx in (1, 2, 3):
            ws.cell(row=row_idx, column=col_idx).alignment = left
        for col_idx in (4, 5, 6, 7):
            ws.cell(row=row_idx, column=col_idx).alignment = center

        # Reference photo hyperlink
        ref = _ref_photo_path(name)
        if ref:
            rc = ws.cell(row=row_idx, column=8, value=ref.name)
            rc.hyperlink = ref.resolve().as_uri()
            rc.font = link_font; rc.alignment = left
        else:
            ws.cell(row=row_idx, column=8, value="(no reference)").font = body_font

        # Embed "In" snapshot
        ws.row_dimensions[row_idx].height = ROW_HEIGHT_PT
        face_path = best_faces.get(first["id"])
        if face_path:
            thumb = _thumbnail_bytes(Path(face_path))
            if thumb:
                xl_img = XLImage(io.BytesIO(thumb))
                xl_img.width = THUMB_PX; xl_img.height = THUMB_PX
                ws.add_image(xl_img, f"{get_column_letter(9)}{row_idx}")

        row_idx += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    if output:
        output_path = Path(output)
    else:
        ts = time.strftime("%H%M%S")
        output_path = REPORTS_DIR / f"attendance_{date}__{ts}.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "ok": True,
        "output_path": str(output_path.resolve()),
        "output_name": output_path.name,
        "rows": len(per_person),
        "people": len(per_person),
        "date": date,
    }


# ---------------------------------------------------------------
# CLI
# ---------------------------------------------------------------
def _cli():
    p = argparse.ArgumentParser(description="Generate daily Excel report")
    p.add_argument("--date", required=True, help="YYYY-MM-DD")
    p.add_argument("--output")
    p.add_argument("--exclude-unknown", action="store_true")
    args = p.parse_args()

    result = generate_report(
        date=args.date, output=args.output,
        include_unknown=not args.exclude_unknown,
    )
    if not result["ok"]:
        sys.exit(result["error"])
    print(f"[done] {result['output_path']}")
    print(f"       {result['rows']} rows  ({result['matched']} matched, "
          f"{result['unknown']} unknown)")


if __name__ == "__main__":
    _cli()
"""Tests for P5: employees CRUD, search, Excel import/export, soft delete."""

from __future__ import annotations

from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from tests.conftest import department_id_by_code


def _login(client: TestClient, user: dict) -> None:
    resp = client.post(
        "/api/auth/login",
        json={"email": user["email"], "password": user["password"]},
    )
    assert resp.status_code == 200, resp.text


def _build_xlsx(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["employee_code", "full_name", "email", "department_code"])
    for r in rows:
        ws.append(
            [
                r.get("employee_code", ""),
                r.get("full_name", ""),
                r.get("email", ""),
                r.get("department_code", ""),
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Import: 5 rows (3 valid, 1 bad dept, 1 duplicate)
# ---------------------------------------------------------------------------


@pytest.mark.usefixtures("clean_employees")
def test_import_5_rows_3_valid_1_bad_dept_1_duplicate(
    client: TestClient, admin_user: dict
) -> None:
    _login(client, admin_user)

    xlsx = _build_xlsx(
        [
            # Row 2 — valid, new
            {"employee_code": "OM0001", "full_name": "Alice Al-Habsi", "email": "alice@example.com", "department_code": "ENG"},
            # Row 3 — valid, new
            {"employee_code": "OM0002", "full_name": "Bob Al-Kindi", "email": "bob@example.com", "department_code": "OPS"},
            # Row 4 — valid, new
            {"employee_code": "OM0003", "full_name": "Carol Al-Busaidi", "email": "carol@example.com", "department_code": "ADM"},
            # Row 5 — bad department
            {"employee_code": "OM0004", "full_name": "Dan Al-Shukaili", "email": "dan@example.com", "department_code": "ZZZ"},
            # Row 6 — duplicate of row 2
            {"employee_code": "OM0001", "full_name": "Alice duplicate", "email": "alice-dup@example.com", "department_code": "ENG"},
        ]
    )

    resp = client.post(
        "/api/employees/import",
        files={"file": ("employees.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["created"] == 3
    assert body["updated"] == 0
    assert len(body["errors"]) == 2

    errors_by_row = {e["row"]: e["message"] for e in body["errors"]}
    assert 5 in errors_by_row
    assert "department" in errors_by_row[5].lower()
    assert 6 in errors_by_row
    assert "duplicate" in errors_by_row[6].lower()

    # Re-import with a corrected version should now be an UPDATE for OM0001.
    xlsx2 = _build_xlsx(
        [
            {"employee_code": "OM0001", "full_name": "Alice Renamed", "email": "alice+new@example.com", "department_code": "ENG"},
        ]
    )
    resp = client.post(
        "/api/employees/import",
        files={"file": ("reimport.xlsx", xlsx2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    body2 = resp.json()
    assert body2["created"] == 0
    assert body2["updated"] == 1
    assert body2["errors"] == []


# ---------------------------------------------------------------------------
# Export round-trip
# ---------------------------------------------------------------------------


@pytest.mark.usefixtures("clean_employees")
def test_export_round_trip_contains_all_expected_columns(
    client: TestClient, admin_user: dict
) -> None:
    _login(client, admin_user)

    client.post(
        "/api/employees/import",
        files={
            "file": (
                "seed.xlsx",
                _build_xlsx(
                    [
                        {"employee_code": "OM0010", "full_name": "Exportable One", "email": "one@example.com", "department_code": "ENG"},
                        {"employee_code": "OM0011", "full_name": "Exportable Two", "email": "two@example.com", "department_code": "OPS"},
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    resp = client.get("/api/employees/export")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "employees.xlsx" in resp.headers["content-disposition"]

    wb = load_workbook(BytesIO(resp.content), read_only=True, data_only=True)
    ws = wb.active
    assert ws is not None

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = list(rows[0])
    # P28.7 added six optional columns to the export. Assert the
    # core six come first (so the column order isn't subtly broken)
    # and the new columns trail in the documented order.
    assert header[:6] == [
        "employee_code",
        "full_name",
        "email",
        "department_code",
        "status",
        "photo_count",
    ]
    assert header[6:] == [
        "designation",
        "phone",
        "reports_to_email",
        "joining_date",
        "relieving_date",
        "deactivation_reason",
    ]

    by_code = {r[0]: r for r in rows[1:]}
    assert "OM0010" in by_code and "OM0011" in by_code
    assert by_code["OM0010"][3] == "ENG"
    assert by_code["OM0010"][4] == "active"
    assert by_code["OM0010"][5] == 0  # photo_count — photos come in P6


# ---------------------------------------------------------------------------
# Search hits
# ---------------------------------------------------------------------------


@pytest.mark.usefixtures("clean_employees")
def test_search_hits_by_code_name_email_and_department(
    client: TestClient, admin_user: dict, admin_engine
) -> None:
    _login(client, admin_user)

    client.post(
        "/api/employees/import",
        files={
            "file": (
                "seed.xlsx",
                _build_xlsx(
                    [
                        {"employee_code": "OM0097", "full_name": "Tariq Al-Shukaili", "email": "tariq@example.com", "department_code": "ENG"},
                        {"employee_code": "OM0098", "full_name": "Fatima Al-Kindi", "email": "fatima@example.com", "department_code": "OPS"},
                        {"employee_code": "OM0099", "full_name": "Layla Al-Busaidi", "email": "layla@example.com", "department_code": "ADM"},
                    ]
                ),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    # By employee_code
    assert _codes(client.get("/api/employees", params={"q": "OM0097"}).json()) == ["OM0097"]
    # By full_name fragment
    assert _codes(client.get("/api/employees", params={"q": "fatima"}).json()) == ["OM0098"]
    # By email fragment
    assert _codes(client.get("/api/employees", params={"q": "layla@"}).json()) == ["OM0099"]
    # By department name
    assert _codes(client.get("/api/employees", params={"q": "engineering"}).json()) == ["OM0097"]

    # Filter by department_id
    ops_id = department_id_by_code(admin_engine, "OPS")
    filtered = client.get("/api/employees", params={"department_id": ops_id}).json()
    assert _codes(filtered) == ["OM0098"]


def _codes(body: dict) -> list[str]:
    return [item["employee_code"] for item in body["items"]]


# ---------------------------------------------------------------------------
# Soft delete hides from default list
# ---------------------------------------------------------------------------


@pytest.mark.usefixtures("clean_employees")
def test_soft_delete_hides_from_list_but_keeps_row(
    client: TestClient, admin_user: dict
) -> None:
    _login(client, admin_user)

    # Create via the POST endpoint (exercises the JSON path too).
    create = client.post(
        "/api/employees",
        json={
            "employee_code": "OM0500",
            "full_name": "To Be Archived",
            "email": "archive@example.com",
            "department_code": "ENG",
        },
    )
    assert create.status_code == 201, create.text
    emp_id = create.json()["id"]

    # Visible by default.
    assert "OM0500" in _codes(client.get("/api/employees").json())

    # Soft delete.
    assert client.delete(f"/api/employees/{emp_id}").status_code == 204

    # Gone from the default list.
    assert "OM0500" not in _codes(client.get("/api/employees").json())

    # But still present when include_inactive=true, and its status is 'inactive'.
    with_inactive = client.get(
        "/api/employees", params={"include_inactive": "true"}
    ).json()
    arch = next((r for r in with_inactive["items"] if r["employee_code"] == "OM0500"), None)
    assert arch is not None
    assert arch["status"] == "inactive"

    # Detail still resolves.
    detail = client.get(f"/api/employees/{emp_id}")
    assert detail.status_code == 200
    assert detail.json()["status"] == "inactive"


# ---------------------------------------------------------------------------
# Role guard: Employee must get 403
# ---------------------------------------------------------------------------


@pytest.mark.usefixtures("clean_employees")
def test_employee_role_is_forbidden(
    client: TestClient, employee_user: dict
) -> None:
    _login(client, employee_user)
    assert client.get("/api/employees").status_code == 403
    assert client.post("/api/employees", json={
        "employee_code": "X", "full_name": "X", "department_code": "ENG"
    }).status_code == 403

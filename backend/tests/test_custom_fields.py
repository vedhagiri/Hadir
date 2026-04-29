"""Tests for v1.0 P12 — custom fields editor.

Covers:

* CRUD on /api/custom-fields (Admin-only field defs, Admin/HR reads).
* Reorder endpoint persists display_order.
* Per-employee value upsert + clear via PATCH; coercion per type
  (text / number / date / select).
* Excel export appends one column per field (ordered by display_order)
  with the per-employee value populated.
* Excel import accepts custom-field columns matched by code; unknown
  columns produce row warnings, not row errors; the standard columns
  still import normally.
* Deleting a field cascades and removes its values for every employee.
* Employee role is forbidden from custom-field defs.

Tests reuse the TENANT_ID=1 (pilot) tenant — the same one
``test_employees.py`` exercises.
"""

from __future__ import annotations

from io import BytesIO
from typing import Iterator

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, select
from sqlalchemy.engine import Engine

from maugood.db import custom_field_values, custom_fields
from tests.conftest import TENANT_ID, department_id_by_code


def _login(client: TestClient, user: dict) -> None:
    resp = client.post(
        "/api/auth/login",
        json={"email": user["email"], "password": user["password"]},
    )
    assert resp.status_code == 200, resp.text


@pytest.fixture
def clean_custom_fields(admin_engine: Engine) -> Iterator[None]:
    with admin_engine.begin() as conn:
        conn.execute(delete(custom_field_values))
        conn.execute(delete(custom_fields))
    yield
    with admin_engine.begin() as conn:
        conn.execute(delete(custom_field_values))
        conn.execute(delete(custom_fields))


def _create_employee(client: TestClient, *, code: str, name: str, dept: str) -> int:
    resp = client.post(
        "/api/employees",
        json={
            "employee_code": code,
            "full_name": name,
            "email": None,
            "department_code": dept,
        },
    )
    assert resp.status_code == 201, resp.text
    return int(resp.json()["id"])


# ---------------------------------------------------------------------------
# Field CRUD
# ---------------------------------------------------------------------------


@pytest.mark.usefixtures("clean_custom_fields", "clean_employees")
def test_create_text_and_select_fields_persist(
    client: TestClient, admin_user: dict
) -> None:
    _login(client, admin_user)

    r1 = client.post(
        "/api/custom-fields",
        json={"name": "Badge Number", "code": "badge_number", "type": "text"},
    )
    assert r1.status_code == 201, r1.text
    assert r1.json()["display_order"] == 0
    assert r1.json()["options"] is None

    r2 = client.post(
        "/api/custom-fields",
        json={
            "name": "Contract Type",
            "code": "contract_type",
            "type": "select",
            "options": ["Permanent", "Contract", "Intern"],
            "required": True,
        },
    )
    assert r2.status_code == 201, r2.text
    assert r2.json()["display_order"] == 1
    assert r2.json()["options"] == ["Permanent", "Contract", "Intern"]
    assert r2.json()["required"] is True

    listing = client.get("/api/custom-fields").json()
    assert [f["code"] for f in listing] == ["badge_number", "contract_type"]


@pytest.mark.usefixtures("clean_custom_fields")
def test_select_field_requires_options(
    client: TestClient, admin_user: dict
) -> None:
    _login(client, admin_user)
    bad = client.post(
        "/api/custom-fields",
        json={"name": "Bad", "code": "bad_select", "type": "select"},
    )
    assert bad.status_code == 422


@pytest.mark.usefixtures("clean_custom_fields")
def test_duplicate_code_returns_409(
    client: TestClient, admin_user: dict
) -> None:
    _login(client, admin_user)
    client.post(
        "/api/custom-fields",
        json={"name": "Badge", "code": "badge", "type": "text"},
    )
    again = client.post(
        "/api/custom-fields",
        json={"name": "Badge2", "code": "badge", "type": "text"},
    )
    assert again.status_code == 409


@pytest.mark.usefixtures("clean_custom_fields")
def test_employee_role_forbidden_on_field_defs(
    client: TestClient, employee_user: dict
) -> None:
    _login(client, employee_user)
    list_resp = client.get("/api/custom-fields")
    assert list_resp.status_code == 403
    create_resp = client.post(
        "/api/custom-fields",
        json={"name": "Nope", "code": "nope", "type": "text"},
    )
    assert create_resp.status_code == 403


@pytest.mark.usefixtures("clean_custom_fields")
def test_reorder_persists(client: TestClient, admin_user: dict) -> None:
    _login(client, admin_user)
    a = client.post(
        "/api/custom-fields",
        json={"name": "A", "code": "a_code", "type": "text"},
    ).json()
    b = client.post(
        "/api/custom-fields",
        json={"name": "B", "code": "b_code", "type": "text"},
    ).json()
    c = client.post(
        "/api/custom-fields",
        json={"name": "C", "code": "c_code", "type": "text"},
    ).json()

    # Initial: a=0, b=1, c=2
    initial = client.get("/api/custom-fields").json()
    assert [f["code"] for f in initial] == ["a_code", "b_code", "c_code"]

    # Move c → 0, a → 1, b → 2
    resp = client.post(
        "/api/custom-fields/reorder",
        json={
            "items": [
                {"id": c["id"], "display_order": 0},
                {"id": a["id"], "display_order": 1},
                {"id": b["id"], "display_order": 2},
            ]
        },
    )
    assert resp.status_code == 200, resp.text
    listed = client.get("/api/custom-fields").json()
    assert [f["code"] for f in listed] == ["c_code", "a_code", "b_code"]


# ---------------------------------------------------------------------------
# Per-employee values
# ---------------------------------------------------------------------------


@pytest.mark.usefixtures("clean_custom_fields", "clean_employees")
def test_value_round_trip_per_type(
    client: TestClient, admin_user: dict
) -> None:
    _login(client, admin_user)

    badge = client.post(
        "/api/custom-fields",
        json={"name": "Badge", "code": "badge", "type": "text"},
    ).json()
    contract = client.post(
        "/api/custom-fields",
        json={
            "name": "Contract",
            "code": "contract",
            "type": "select",
            "options": ["Permanent", "Contract"],
        },
    ).json()
    yos = client.post(
        "/api/custom-fields",
        json={"name": "Years of Service", "code": "yos", "type": "number"},
    ).json()
    start = client.post(
        "/api/custom-fields",
        json={"name": "Start Date", "code": "start_date", "type": "date"},
    ).json()

    emp_id = _create_employee(client, code="OM0501", name="Test Emp", dept="ENG")

    patch = client.patch(
        f"/api/employees/{emp_id}/custom-fields",
        json={
            "items": [
                {"field_id": badge["id"], "value": "B-1234"},
                {"field_id": contract["id"], "value": "Permanent"},
                {"field_id": yos["id"], "value": 5},
                {"field_id": start["id"], "value": "2024-01-15"},
            ]
        },
    )
    assert patch.status_code == 200, patch.text

    by_code = {v["code"]: v for v in patch.json()}
    assert by_code["badge"]["value"] == "B-1234"
    assert by_code["contract"]["value"] == "Permanent"
    assert by_code["yos"]["value"] == 5
    assert by_code["start_date"]["value"] == "2024-01-15"


@pytest.mark.usefixtures("clean_custom_fields", "clean_employees")
def test_invalid_select_value_rejected_422(
    client: TestClient, admin_user: dict
) -> None:
    _login(client, admin_user)
    field = client.post(
        "/api/custom-fields",
        json={
            "name": "Contract",
            "code": "contract",
            "type": "select",
            "options": ["Permanent", "Contract"],
        },
    ).json()
    emp_id = _create_employee(client, code="OM0502", name="X", dept="ENG")

    bad = client.patch(
        f"/api/employees/{emp_id}/custom-fields",
        json={"items": [{"field_id": field["id"], "value": "Freelance"}]},
    )
    assert bad.status_code == 422


@pytest.mark.usefixtures("clean_custom_fields", "clean_employees")
def test_value_clear_via_null(client: TestClient, admin_user: dict) -> None:
    _login(client, admin_user)
    field = client.post(
        "/api/custom-fields",
        json={"name": "Note", "code": "note", "type": "text"},
    ).json()
    emp_id = _create_employee(client, code="OM0503", name="Y", dept="ENG")

    client.patch(
        f"/api/employees/{emp_id}/custom-fields",
        json={"items": [{"field_id": field["id"], "value": "hello"}]},
    )
    listing = client.get(f"/api/employees/{emp_id}/custom-fields").json()
    assert listing[0]["value"] == "hello"

    client.patch(
        f"/api/employees/{emp_id}/custom-fields",
        json={"items": [{"field_id": field["id"], "value": None}]},
    )
    listing2 = client.get(f"/api/employees/{emp_id}/custom-fields").json()
    assert listing2[0]["value"] is None


# ---------------------------------------------------------------------------
# Excel round-trip
# ---------------------------------------------------------------------------


@pytest.mark.usefixtures("clean_custom_fields", "clean_employees")
def test_export_includes_custom_field_columns(
    client: TestClient, admin_user: dict
) -> None:
    _login(client, admin_user)

    badge = client.post(
        "/api/custom-fields",
        json={"name": "Badge", "code": "badge", "type": "text"},
    ).json()
    contract = client.post(
        "/api/custom-fields",
        json={
            "name": "Contract",
            "code": "contract",
            "type": "select",
            "options": ["Permanent", "Contract"],
        },
    ).json()

    emp_id = _create_employee(client, code="OM0701", name="Z", dept="ENG")
    client.patch(
        f"/api/employees/{emp_id}/custom-fields",
        json={
            "items": [
                {"field_id": badge["id"], "value": "B-7777"},
                {"field_id": contract["id"], "value": "Permanent"},
            ]
        },
    )

    resp = client.get("/api/employees/export")
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.content), data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "badge" in headers
    assert "contract" in headers
    # Data row carries the values
    badge_idx = headers.index("badge")
    contract_idx = headers.index("contract")
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    om = next(r for r in rows if r[0] == "OM0701")
    assert om[badge_idx] == "B-7777"
    assert om[contract_idx] == "Permanent"


def _build_xlsx_with_extras(
    rows: list[dict], extra_headers: list[str]
) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(
        ["employee_code", "full_name", "email", "department_code", *extra_headers]
    )
    for r in rows:
        ws.append(
            [
                r["employee_code"],
                r["full_name"],
                r.get("email", ""),
                r["department_code"],
                *(r.get(h, "") for h in extra_headers),
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.usefixtures("clean_custom_fields", "clean_employees")
def test_import_applies_known_codes_and_warns_on_unknown(
    client: TestClient, admin_user: dict
) -> None:
    _login(client, admin_user)

    badge = client.post(
        "/api/custom-fields",
        json={"name": "Badge", "code": "badge", "type": "text"},
    ).json()

    xlsx = _build_xlsx_with_extras(
        rows=[
            {
                "employee_code": "OM0801",
                "full_name": "Imp One",
                "email": "imp1@test.maugood",
                "department_code": "ENG",
                "badge": "B-100",
                "ghost_field": "ignored",
            },
            {
                "employee_code": "OM0802",
                "full_name": "Imp Two",
                "email": "imp2@test.maugood",
                "department_code": "OPS",
                "badge": "B-200",
                "ghost_field": "also ignored",
            },
        ],
        extra_headers=["badge", "ghost_field"],
    )

    resp = client.post(
        "/api/employees/import",
        files={
            "file": (
                "employees.xlsx",
                xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["created"] == 2
    assert body["errors"] == []
    # ``ghost_field`` produced one warning per row
    warning_messages = [w["message"] for w in body["warnings"]]
    assert sum("ghost_field" in m for m in warning_messages) == 2

    # Badge values stuck against both employees.
    listing = client.get("/api/employees?include_inactive=true").json()
    by_code = {e["employee_code"]: e for e in listing["items"]}
    om1_id = by_code["OM0801"]["id"]
    om2_id = by_code["OM0802"]["id"]
    om1_values = client.get(f"/api/employees/{om1_id}/custom-fields").json()
    om2_values = client.get(f"/api/employees/{om2_id}/custom-fields").json()
    assert next(v for v in om1_values if v["code"] == "badge")["value"] == "B-100"
    assert next(v for v in om2_values if v["code"] == "badge")["value"] == "B-200"


@pytest.mark.usefixtures("clean_custom_fields", "clean_employees")
def test_import_coerce_failure_warns_and_skips_cell(
    client: TestClient, admin_user: dict
) -> None:
    _login(client, admin_user)
    client.post(
        "/api/custom-fields",
        json={"name": "Years", "code": "years", "type": "number"},
    )

    xlsx = _build_xlsx_with_extras(
        rows=[
            {
                "employee_code": "OM0901",
                "full_name": "Bad Number",
                "email": "bn@test.maugood",
                "department_code": "ENG",
                "years": "not-a-number",
            }
        ],
        extra_headers=["years"],
    )
    resp = client.post(
        "/api/employees/import",
        files={
            "file": (
                "employees.xlsx",
                xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    body = resp.json()
    assert body["created"] == 1  # the standard row still imported
    assert body["errors"] == []
    assert any("number" in w["message"].lower() for w in body["warnings"])

    # And no value was persisted for the bad cell.
    listing = client.get("/api/employees?include_inactive=true").json()
    om_id = next(e["id"] for e in listing["items"] if e["employee_code"] == "OM0901")
    values = client.get(f"/api/employees/{om_id}/custom-fields").json()
    assert next(v for v in values if v["code"] == "years")["value"] is None


# ---------------------------------------------------------------------------
# Cascade delete
# ---------------------------------------------------------------------------


@pytest.mark.usefixtures("clean_custom_fields", "clean_employees")
def test_delete_field_cascades_values(
    client: TestClient, admin_user: dict, admin_engine: Engine
) -> None:
    _login(client, admin_user)
    field = client.post(
        "/api/custom-fields",
        json={"name": "Note", "code": "note", "type": "text"},
    ).json()
    emp_id = _create_employee(client, code="OM1001", name="Cas", dept="ENG")
    client.patch(
        f"/api/employees/{emp_id}/custom-fields",
        json={"items": [{"field_id": field["id"], "value": "remember me"}]},
    )

    # Confirm the value row exists.
    with admin_engine.begin() as conn:
        before = conn.execute(
            select(custom_field_values.c.id).where(
                custom_field_values.c.field_id == field["id"]
            )
        ).all()
    assert len(before) == 1

    resp = client.delete(f"/api/custom-fields/{field['id']}")
    assert resp.status_code == 204

    with admin_engine.begin() as conn:
        after = conn.execute(
            select(custom_field_values.c.id).where(
                custom_field_values.c.field_id == field["id"]
            )
        ).all()
    assert after == []

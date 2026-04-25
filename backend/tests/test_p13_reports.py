"""Tests for P13 — Excel report endpoint + dev-only smoke endpoints."""

from __future__ import annotations

from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import delete, insert, select

from hadir.cameras import rtsp as rtsp_io
from hadir.config import get_settings
from hadir.db import (
    attendance_records,
    cameras,
    detection_events,
    employees,
    shift_policies,
    user_departments,
)


def _login(client: TestClient, user: dict) -> None:
    resp = client.post(
        "/api/auth/login",
        json={"email": user["email"], "password": user["password"]},
    )
    assert resp.status_code == 200, resp.text


def _seed_attendance(admin_engine) -> dict:  # type: ignore[no-untyped-def]
    """Seed two employees in two departments and three attendance rows."""

    today = date.today()

    with admin_engine.begin() as conn:
        conn.execute(delete(attendance_records).where(attendance_records.c.tenant_id == 1))
        conn.execute(delete(detection_events).where(detection_events.c.tenant_id == 1))
        conn.execute(delete(cameras).where(cameras.c.tenant_id == 1))
        conn.execute(delete(employees).where(employees.c.tenant_id == 1))

        eng_id = conn.execute(
            insert(employees)
            .values(
                tenant_id=1,
                employee_code="P13-ENG",
                full_name="Tariq Al-Shukaili",
                email="tariq@p13.example",
                department_id=1,  # ENG
                status="active",
            )
            .returning(employees.c.id)
        ).scalar_one()
        ops_id = conn.execute(
            insert(employees)
            .values(
                tenant_id=1,
                employee_code="P13-OPS",
                full_name="Fatima Al-Kindi",
                email="fatima@p13.example",
                department_id=2,  # OPS
                status="active",
            )
            .returning(employees.c.id)
        ).scalar_one()

        # Use the seeded pilot policy (id=1) — already in shift_policies.
        policy_id = conn.execute(
            select(shift_policies.c.id)
            .where(shift_policies.c.tenant_id == 1)
            .order_by(shift_policies.c.id.asc())
            .limit(1)
        ).scalar_one()

        # Today + yesterday rows for ENG, today only for OPS.
        for emp_id, the_date, in_t, out_t, late, ot in (
            (eng_id, today, time(7, 28), time(15, 36), False, 8),
            (eng_id, today - timedelta(days=1), time(7, 31), time(15, 30), False, 0),
            (ops_id, today, time(7, 50), time(15, 5), True, 0),
        ):
            total = (out_t.hour * 60 + out_t.minute) - (in_t.hour * 60 + in_t.minute)
            conn.execute(
                insert(attendance_records).values(
                    tenant_id=1,
                    employee_id=emp_id,
                    date=the_date,
                    in_time=in_t,
                    out_time=out_t,
                    total_minutes=total,
                    policy_id=policy_id,
                    late=late,
                    early_out=False,
                    short_hours=total < 480,
                    absent=False,
                    overtime_minutes=ot,
                )
            )

    return {"eng_id": int(eng_id), "ops_id": int(ops_id), "today": today}


@pytest.fixture
def seeded_attendance(admin_engine):  # type: ignore[no-untyped-def]
    info = _seed_attendance(admin_engine)
    try:
        yield info
    finally:
        with admin_engine.begin() as conn:
            conn.execute(delete(attendance_records).where(attendance_records.c.tenant_id == 1))
            conn.execute(delete(detection_events).where(detection_events.c.tenant_id == 1))
            conn.execute(delete(cameras).where(cameras.c.tenant_id == 1))
            conn.execute(delete(employees).where(employees.c.tenant_id == 1))


# ---------------------------------------------------------------------------
# /api/reports/attendance.xlsx — round trip
# ---------------------------------------------------------------------------


def test_report_round_trip_contains_expected_rows(
    client: TestClient, admin_user: dict, seeded_attendance
) -> None:
    _login(client, admin_user)
    today = seeded_attendance["today"]
    resp = client.post(
        "/api/reports/attendance.xlsx",
        json={
            "start": (today - timedelta(days=1)).isoformat(),
            "end": today.isoformat(),
        },
    )
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attendance_" in resp.headers["content-disposition"]

    wb = load_workbook(BytesIO(resp.content), read_only=True, data_only=True)
    # Sheets named by ISO week. Could be one or two depending on whether
    # today + yesterday cross a week boundary; the row content is what
    # we actually care about.
    all_data_rows: list[tuple] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        assert header[:3] == ("employee_code", "name", "date")
        for r in rows[1:]:
            all_data_rows.append(r)
    wb.close()

    codes = {r[0] for r in all_data_rows}
    assert {"P13-ENG", "P13-OPS"} <= codes
    # ENG should have 2 rows (today + yesterday); OPS only today.
    eng_rows = [r for r in all_data_rows if r[0] == "P13-ENG"]
    ops_rows = [r for r in all_data_rows if r[0] == "P13-OPS"]
    assert len(eng_rows) == 2
    assert len(ops_rows) == 1


def test_report_filters_to_one_department(
    client: TestClient, admin_user: dict, seeded_attendance
) -> None:
    _login(client, admin_user)
    today = seeded_attendance["today"]
    resp = client.post(
        "/api/reports/attendance.xlsx",
        json={
            "start": today.isoformat(),
            "end": today.isoformat(),
            "department_id": 1,  # ENG only
        },
    )
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.content), read_only=True, data_only=True)
    codes: set[str] = set()
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for r in list(ws.iter_rows(values_only=True))[1:]:
            codes.add(r[0])
    wb.close()
    assert codes == {"P13-ENG"}


def test_report_403_for_employee_role(
    client: TestClient, employee_user: dict, seeded_attendance
) -> None:
    _login(client, employee_user)
    today = seeded_attendance["today"]
    resp = client.post(
        "/api/reports/attendance.xlsx",
        json={"start": today.isoformat(), "end": today.isoformat()},
    )
    assert resp.status_code == 403


def test_manager_scoped_to_assigned_departments(
    client: TestClient, admin_user: dict, seeded_attendance, admin_engine
) -> None:
    """A Manager assigned to ENG (id=1) only sees ENG rows in the report.

    We borrow the ``admin_user`` fixture and reassign their roles +
    department membership for this test, then restore them.
    """

    from sqlalchemy import insert as _insert
    from hadir.db import roles, user_roles

    with admin_engine.begin() as conn:
        # Drop the Admin role and grant Manager.
        conn.execute(
            delete(user_roles).where(user_roles.c.user_id == admin_user["id"])
        )
        manager_role_id = conn.execute(
            select(roles.c.id).where(
                roles.c.tenant_id == 1, roles.c.code == "Manager"
            )
        ).scalar_one()
        conn.execute(
            _insert(user_roles).values(
                user_id=admin_user["id"],
                role_id=manager_role_id,
                tenant_id=1,
            )
        )
        # Assign to ENG only.
        conn.execute(
            delete(user_departments).where(user_departments.c.user_id == admin_user["id"])
        )
        conn.execute(
            _insert(user_departments).values(
                user_id=admin_user["id"], department_id=1, tenant_id=1
            )
        )

    try:
        _login(client, admin_user)
        today = seeded_attendance["today"]

        # No filter → manager auto-scoped to ENG only.
        resp = client.post(
            "/api/reports/attendance.xlsx",
            json={"start": today.isoformat(), "end": today.isoformat()},
        )
        assert resp.status_code == 200
        wb = load_workbook(BytesIO(resp.content), read_only=True, data_only=True)
        codes: set[str] = set()
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for r in list(ws.iter_rows(values_only=True))[1:]:
                codes.add(r[0])
        wb.close()
        assert codes == {"P13-ENG"}, f"expected ENG only, got {codes}"

        # Filtering by OPS (id=2) — outside their set → 403.
        resp = client.post(
            "/api/reports/attendance.xlsx",
            json={
                "start": today.isoformat(),
                "end": today.isoformat(),
                "department_id": 2,
            },
        )
        assert resp.status_code == 403
    finally:
        # Restore Admin role + drop dept assignment.
        with admin_engine.begin() as conn:
            conn.execute(delete(user_roles).where(user_roles.c.user_id == admin_user["id"]))
            conn.execute(
                delete(user_departments).where(user_departments.c.user_id == admin_user["id"])
            )


def test_report_rejects_invalid_date_range(
    client: TestClient, admin_user: dict
) -> None:
    _login(client, admin_user)
    resp = client.post(
        "/api/reports/attendance.xlsx",
        json={"start": "2026-04-25", "end": "2026-04-20"},
    )
    assert resp.status_code == 400


def test_report_rejects_excessive_range(
    client: TestClient, admin_user: dict
) -> None:
    _login(client, admin_user)
    resp = client.post(
        "/api/reports/attendance.xlsx",
        json={"start": "2024-01-01", "end": "2026-12-31"},
    )
    assert resp.status_code == 400


# ---------------------------------------------------------------------------
# Dev-only test endpoints
# ---------------------------------------------------------------------------


def test_dev_only_endpoints_are_mounted_when_env_is_dev(
    client: TestClient, admin_user: dict, seeded_attendance
) -> None:
    """Smoke that POST /api/_test/seed_detection works (HADIR_ENV=dev in this stack)."""

    assert get_settings().env == "dev"
    _login(client, admin_user)
    resp = client.post(
        "/api/_test/seed_detection",
        json={"employee_code": "P13-ENG", "minutes_offset": -1},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["employee_id"] == seeded_attendance["eng_id"]
    assert body["detection_event_id"] > 0


def test_dev_only_recompute_endpoint(
    client: TestClient, admin_user: dict, seeded_attendance
) -> None:
    _login(client, admin_user)
    resp = client.post("/api/_test/recompute_attendance")
    assert resp.status_code == 200
    body = resp.json()
    # We have 2 active employees seeded; the recompute upserts both.
    assert body["upserted"] >= 2


def test_dev_endpoints_403_for_employee(
    client: TestClient, employee_user: dict, seeded_attendance
) -> None:
    _login(client, employee_user)
    assert (
        client.post(
            "/api/_test/seed_detection",
            json={"employee_code": "P13-ENG"},
        ).status_code
        == 403
    )
    assert client.post("/api/_test/recompute_attendance").status_code == 403

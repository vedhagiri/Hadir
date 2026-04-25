"""End-to-end smoke for v1.0 P12 — custom fields editor.

Logs in as the seeded pilot Admin against the ``main`` tenant
(Pydantic ``EmailStr`` rejects reserved TLDs like ``.test``, which the
v1.0 seeded ``admin@omran.test`` accounts use; the pilot's
``admin@pilot.hadir`` is the working credential set), defines two
custom fields (Badge Number text + Contract Type select), creates a
fresh employee, sets values for both fields, exports the employee
XLSX, and asserts the columns and per-employee values are present in
the file.

Run inside the backend container:

    docker compose exec -e HADIR_SMOKE_PASSWORD='…' backend \\
        python -m scripts.v1_p12_smoke

Cleans up after itself so re-running starts clean.
"""

from __future__ import annotations

import io
import os
import sys

import httpx
from openpyxl import load_workbook


BASE = "http://localhost:8000"
ADMIN_EMAIL = "admin@pilot.hadir"
ADMIN_PASSWORD = os.environ.get("HADIR_SMOKE_PASSWORD", "")
EMPLOYEE_CODE = "OM-P12-SMOKE"


def main() -> int:
    if not ADMIN_PASSWORD:
        print(
            "[p12] set HADIR_SMOKE_PASSWORD to the seeded pilot Admin password",
            file=sys.stderr,
        )
        return 1
    with httpx.Client(base_url=BASE, follow_redirects=False, timeout=20) as c:
        login = c.post(
            "/api/auth/login",
            json={
                "email": ADMIN_EMAIL,
                "password": ADMIN_PASSWORD,
            },
        )
        login.raise_for_status()
        print(f"[p12] login OK as {ADMIN_EMAIL} on main tenant")

        # 1) Reset any lingering smoke state.
        existing = c.get("/api/custom-fields").json()
        for f in existing:
            if f["code"] in ("badge_number", "contract_type"):
                c.delete(f"/api/custom-fields/{f['id']}").raise_for_status()
                print(f"[p12] dropped lingering field {f['code']}")
        emp_list = c.get(
            "/api/employees", params={"q": EMPLOYEE_CODE, "include_inactive": "true"}
        ).json()
        for e in emp_list["items"]:
            if e["employee_code"] == EMPLOYEE_CODE:
                c.delete(f"/api/employees/{e['id']}").raise_for_status()

        # 2) Define two fields.
        badge = c.post(
            "/api/custom-fields",
            json={
                "name": "Badge Number",
                "code": "badge_number",
                "type": "text",
            },
        )
        badge.raise_for_status()
        badge_obj = badge.json()
        print(
            f"[p12] created badge_number id={badge_obj['id']} "
            f"display_order={badge_obj['display_order']}"
        )

        contract = c.post(
            "/api/custom-fields",
            json={
                "name": "Contract Type",
                "code": "contract_type",
                "type": "select",
                "options": ["Permanent", "Contract", "Intern"],
                "required": True,
            },
        )
        contract.raise_for_status()
        contract_obj = contract.json()
        print(
            f"[p12] created contract_type id={contract_obj['id']} "
            f"options={contract_obj['options']}"
        )

        # 3) Create an employee in ENG and set values.
        emp = c.post(
            "/api/employees",
            json={
                "employee_code": EMPLOYEE_CODE,
                "full_name": "P12 Smoke Employee",
                "email": None,
                "department_code": "ENG",
            },
        )
        emp.raise_for_status()
        emp_obj = emp.json()
        emp_id = emp_obj["id"]
        print(f"[p12] created employee id={emp_id} code={EMPLOYEE_CODE}")

        patch = c.patch(
            f"/api/employees/{emp_id}/custom-fields",
            json={
                "items": [
                    {"field_id": badge_obj["id"], "value": "B-9001"},
                    {"field_id": contract_obj["id"], "value": "Permanent"},
                ]
            },
        )
        patch.raise_for_status()
        by_code = {v["code"]: v["value"] for v in patch.json()}
        print(f"[p12] PATCH values → {by_code}")

        # 4) Round-trip via GET.
        got = c.get(f"/api/employees/{emp_id}/custom-fields").json()
        got_map = {v["code"]: v["value"] for v in got}
        assert got_map == {"badge_number": "B-9001", "contract_type": "Permanent"}, got_map
        print(f"[p12] GET values match: {got_map}")

        # 5) Excel export contains the columns + values.
        export = c.get("/api/employees/export")
        export.raise_for_status()
        wb = load_workbook(io.BytesIO(export.content), data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "badge_number" in headers, headers
        assert "contract_type" in headers, headers
        badge_idx = headers.index("badge_number")
        contract_idx = headers.index("contract_type")
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        smoke_row = next(r for r in rows if r[0] == EMPLOYEE_CODE)
        assert smoke_row[badge_idx] == "B-9001", smoke_row
        assert smoke_row[contract_idx] == "Permanent", smoke_row
        print(
            f"[p12] export contains '{EMPLOYEE_CODE}' with "
            f"badge_number={smoke_row[badge_idx]!r} "
            f"contract_type={smoke_row[contract_idx]!r}"
        )

        # 6) Cleanup so re-running this script is idempotent.
        c.delete(f"/api/employees/{emp_id}").raise_for_status()
        c.delete(f"/api/custom-fields/{badge_obj['id']}").raise_for_status()
        c.delete(f"/api/custom-fields/{contract_obj['id']}").raise_for_status()
        print("[p12] cleanup complete")

    print("[p12] OK")
    return 0


if __name__ == "__main__":
    sys.exit(main())

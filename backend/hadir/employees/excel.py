"""Excel import + export using openpyxl.

Kept tiny and dependency-light: no pandas, no abstraction framework. The
import parser normalises headers (``lower``, ``strip``, space→underscore)
so operator-friendly variations like "Employee Code" still match the
canonical ``employee_code`` column.

Row numbers reported in errors are **Excel row numbers** (1-indexed,
header is row 1, first data row is row 2). That's what a human opening
the file in Excel will see.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Iterator, Optional

from openpyxl import Workbook, load_workbook

from hadir.employees.repository import EmployeeRow

REQUIRED_COLUMNS: tuple[str, ...] = (
    "employee_code",
    "full_name",
    "email",
    "department_code",
)

EXPORT_COLUMNS: tuple[str, ...] = (
    "employee_code",
    "full_name",
    "email",
    "department_code",
    "status",
    "photo_count",
)


class ImportParseError(Exception):
    """Raised when the file itself is invalid (missing headers, not XLSX)."""


@dataclass(frozen=True, slots=True)
class ImportRow:
    """One parsed data row. ``excel_row`` is 1-indexed from the top of the file."""

    excel_row: int
    employee_code: str
    full_name: str
    email: Optional[str]
    department_code: str


def _normalise_header(raw: object) -> str:
    if raw is None:
        return ""
    return str(raw).strip().lower().replace(" ", "_")


def _cell_str(value: object) -> str:
    """Coerce a cell value to a trimmed string; empty for None/whitespace."""

    if value is None:
        return ""
    text = str(value).strip()
    return text


def parse_import(stream: BytesIO) -> Iterator[ImportRow]:
    """Yield ``ImportRow`` for each non-empty data row.

    Raises ``ImportParseError`` if the workbook can't be opened or if any
    required column is missing.
    """

    try:
        # read_only so we don't load the whole grid into memory for a
        # thousand-row file — openpyxl streams rows from the archive.
        wb = load_workbook(stream, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — openpyxl raises a zoo of types
        raise ImportParseError(f"could not read workbook: {exc}") from exc

    try:
        ws = wb.active
        if ws is None:
            raise ImportParseError("workbook has no active sheet")

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            raise ImportParseError("workbook is empty")

        headers = [_normalise_header(c) for c in header_row]
        missing = [c for c in REQUIRED_COLUMNS if c not in headers]
        if missing:
            raise ImportParseError(
                f"missing required column(s): {', '.join(missing)}"
            )

        idx = {name: headers.index(name) for name in REQUIRED_COLUMNS}

        for excel_row, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            # Skip fully-empty rows so trailing blank lines don't generate
            # a dozen "missing employee_code" errors.
            if all(v is None or _cell_str(v) == "" for v in row):
                continue

            code = _cell_str(row[idx["employee_code"]]) if idx["employee_code"] < len(row) else ""
            name = _cell_str(row[idx["full_name"]]) if idx["full_name"] < len(row) else ""
            email_raw = _cell_str(row[idx["email"]]) if idx["email"] < len(row) else ""
            dept_code = _cell_str(row[idx["department_code"]]) if idx["department_code"] < len(row) else ""

            yield ImportRow(
                excel_row=excel_row,
                employee_code=code,
                full_name=name,
                email=email_raw or None,
                department_code=dept_code,
            )
    finally:
        wb.close()


def build_export(rows: list[EmployeeRow]) -> BytesIO:
    """Produce an XLSX in-memory with the EXPORT_COLUMNS header + one row each."""

    wb = Workbook()
    ws = wb.active
    assert ws is not None  # openpyxl always returns an active sheet on a fresh WB
    ws.title = "Employees"

    ws.append(list(EXPORT_COLUMNS))
    for row in rows:
        ws.append(
            [
                row.employee_code,
                row.full_name,
                row.email or "",
                row.department_code,
                row.status,
                row.photo_count,
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

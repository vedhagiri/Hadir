"""Attendance Calendar endpoints (P28.6).

Four routes — all tenant-scoped; role scope per endpoint:

* ``GET /api/attendance/calendar/company?month=YYYY-MM`` — Admin/HR
  see tenant-wide; Manager sees their visible employee union;
  Employee 403 (they have no team view).
* ``GET /api/attendance/calendar/person/{employee_id}?month=YYYY-MM``
  — Admin/HR for any employee; Manager only inside their visible
  set; Employee only for self. Out-of-scope ``employee_id`` returns
  **404**, not 403 — 403 leaks role/tenant information.
* ``GET /api/attendance/calendar/day/{employee_id}/{date}`` — same
  scope rules as person.
* ``GET /api/attendance/calendar/export?employee_id=...&month=YYYY-MM``
  — XLSX. Single-employee per-person view; without
  ``employee_id`` returns the company aggregate. Same role gates.
"""

from __future__ import annotations

import io
import logging
from dataclasses import asdict
from datetime import date as date_type, datetime, timezone
from typing import Annotated, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import and_, func, select

from hadir.attendance.repository import load_tenant_settings, local_tz_for
from hadir.attendance_calendar import queries
from hadir.auth.audit import write_audit
from hadir.auth.dependencies import CurrentUser, current_user
from hadir.db import employees, get_engine
from hadir.manager_assignments.repository import (
    get_manager_visible_employee_ids,
)
from hadir.tenants.scope import TenantScope

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/api/attendance/calendar", tags=["attendance-calendar"]
)


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------


class CompanyDayOut(BaseModel):
    date: date_type
    present_count: int
    late_count: int
    absent_count: int
    leave_count: int
    active_employees: int
    is_weekend: bool
    is_holiday: bool
    holiday_name: Optional[str] = None
    percent_present: int


class CompanyMonthOut(BaseModel):
    month: str
    days: list[CompanyDayOut]


class PersonDayOut(BaseModel):
    date: date_type
    status: str
    in_time: Optional[str] = None
    out_time: Optional[str] = None
    total_minutes: Optional[int] = None
    overtime_minutes: int = 0
    policy_name: Optional[str] = None
    is_weekend: bool
    is_holiday: bool
    holiday_name: Optional[str] = None
    leave_name: Optional[str] = None


class PersonMonthOut(BaseModel):
    month: str
    employee_id: int
    employee_code: str
    full_name: str
    days: list[PersonDayOut]


class TimelineIntervalOut(BaseModel):
    start: str
    end: str


class EvidenceOut(BaseModel):
    detection_event_id: int
    captured_at: str
    camera_code: str
    confidence: Optional[float] = None
    crop_url: str


class DayDetailOut(BaseModel):
    employee_id: int
    employee_code: str
    full_name: str
    department_name: str
    date: date_type
    status: str
    in_time: Optional[str] = None
    out_time: Optional[str] = None
    total_minutes: Optional[int] = None
    overtime_minutes: int
    policy_name: Optional[str] = None
    policy_description: Optional[str] = None
    policy_scope: str
    timeline: list[TimelineIntervalOut]
    evidence: list[EvidenceOut]
    is_weekend: bool
    is_holiday: bool
    holiday_name: Optional[str] = None
    leave_name: Optional[str] = None


# ---------------------------------------------------------------------------
# Role helpers
# ---------------------------------------------------------------------------


def _employee_row_id_for(user: CurrentUser) -> Optional[int]:
    """Map the logged-in user to an ``employees.id`` by lower-cased email
    — the same convention the existing attendance router uses (pilot
    has no explicit ``user_id → employee_id`` join table)."""

    with get_engine().begin() as conn:
        row = conn.execute(
            select(employees.c.id).where(
                employees.c.tenant_id == user.tenant_id,
                func.lower(employees.c.email) == (user.email or "").lower(),
            )
        ).first()
    return int(row.id) if row is not None else None


def _is_admin_like(user: CurrentUser) -> bool:
    return "Admin" in user.roles or "HR" in user.roles


def _is_manager(user: CurrentUser) -> bool:
    return "Manager" in user.roles


def _resolve_visible_employees(
    user: CurrentUser, scope: TenantScope
) -> Optional[list[int]]:
    """Return the role-scoped employee_id allow-list, or ``None`` for
    Admin/HR (= no narrowing).

    Manager → union of department membership + manager_assignments.
    Employee → just the employee row that maps to their email.
    """

    if _is_admin_like(user):
        return None
    if _is_manager(user):
        with get_engine().begin() as conn:
            visible = get_manager_visible_employee_ids(
                conn, scope, manager_user_id=user.id
            )
        return sorted(int(x) for x in visible)
    # Employee
    eid = _employee_row_id_for(user)
    return [eid] if eid is not None else []


def _check_can_view_employee(
    user: CurrentUser, scope: TenantScope, employee_id: int
) -> None:
    """Raise 404 if the user can't see this employee. Always 404,
    never 403 — 403 leaks "this id exists in another scope".
    """

    # First confirm the employee row even exists in this tenant. A
    # cross-tenant guess returns 404 — never reveals an inaisys row
    # to an mts_demo Admin.
    with get_engine().begin() as conn:
        row = conn.execute(
            select(employees.c.id).where(
                employees.c.tenant_id == scope.tenant_id,
                employees.c.id == employee_id,
            )
        ).first()
    if row is None:
        raise HTTPException(status_code=404, detail="employee not found")

    if _is_admin_like(user):
        return
    visible = _resolve_visible_employees(user, scope)
    if visible is None or employee_id in visible:
        return
    raise HTTPException(status_code=404, detail="employee not found")


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


@router.get("/company", response_model=CompanyMonthOut)
def get_company_calendar(
    user: Annotated[CurrentUser, Depends(current_user)],
    month: Annotated[
        str, Query(description="YYYY-MM month bounds for the calendar.")
    ],
) -> CompanyMonthOut:
    if not (_is_admin_like(user) or _is_manager(user)):
        # Employee role: the UI hides this view; the API hardens the
        # gate. 403 here is intentional — Employees can't see the
        # whole-company aggregate.
        raise HTTPException(status_code=403, detail="forbidden")

    try:
        month_start, month_end = queries.parse_month(month)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    scope = TenantScope(tenant_id=user.tenant_id)
    employee_ids = _resolve_visible_employees(user, scope)

    with get_engine().begin() as conn:
        days = queries.company_view(
            conn,
            scope,
            month_start=month_start,
            month_end=month_end,
            employee_ids=employee_ids,
        )

    return CompanyMonthOut(
        month=month,
        days=[CompanyDayOut(**asdict(d)) for d in days],
    )


@router.get(
    "/person/{employee_id}", response_model=PersonMonthOut
)
def get_person_calendar(
    employee_id: int,
    user: Annotated[CurrentUser, Depends(current_user)],
    month: Annotated[str, Query(...)],
) -> PersonMonthOut:
    try:
        month_start, month_end = queries.parse_month(month)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    scope = TenantScope(tenant_id=user.tenant_id)
    _check_can_view_employee(user, scope, employee_id)

    with get_engine().begin() as conn:
        emp_row = conn.execute(
            select(
                employees.c.id,
                employees.c.employee_code,
                employees.c.full_name,
            ).where(
                employees.c.tenant_id == scope.tenant_id,
                employees.c.id == employee_id,
            )
        ).first()
        if emp_row is None:
            raise HTTPException(status_code=404, detail="employee not found")

        settings = load_tenant_settings(conn, scope)
        today_local = datetime.now(timezone.utc).astimezone(
            local_tz_for(settings)
        ).date()

        days = queries.person_view(
            conn,
            scope,
            employee_id=employee_id,
            month_start=month_start,
            month_end=month_end,
            today_local=today_local,
        )

    return PersonMonthOut(
        month=month,
        employee_id=int(emp_row.id),
        employee_code=str(emp_row.employee_code),
        full_name=str(emp_row.full_name),
        days=[PersonDayOut(**asdict(d)) for d in days],
    )


@router.get(
    "/day/{employee_id}/{the_date}", response_model=DayDetailOut
)
def get_day_detail(
    employee_id: int,
    the_date: date_type,
    user: Annotated[CurrentUser, Depends(current_user)],
) -> DayDetailOut:
    scope = TenantScope(tenant_id=user.tenant_id)
    _check_can_view_employee(user, scope, employee_id)

    with get_engine().begin() as conn:
        detail = queries.fetch_day_detail(
            conn,
            scope,
            employee_id=employee_id,
            the_date=the_date,
        )
    if detail is None:
        # Already filtered by tenant in _check_can_view_employee, but
        # belt-and-braces.
        raise HTTPException(status_code=404, detail="not found")

    # Audit on read — same red line as the existing
    # ``detection_event.crop_viewed``: if a crop is being looked at,
    # we record who looked. The list endpoints don't audit (would
    # flood at one row per nav).
    with get_engine().begin() as conn:
        write_audit(
            conn,
            tenant_id=scope.tenant_id,
            actor_user_id=user.id if user.id > 0 else None,
            action="attendance_calendar.day_viewed",
            entity_type="employee",
            entity_id=str(employee_id),
            after={"date": the_date.isoformat()},
        )

    return DayDetailOut(
        employee_id=detail.employee_id,
        employee_code=detail.employee_code,
        full_name=detail.full_name,
        department_name=detail.department_name,
        date=detail.date,
        status=detail.status,
        in_time=detail.in_time,
        out_time=detail.out_time,
        total_minutes=detail.total_minutes,
        overtime_minutes=detail.overtime_minutes,
        policy_name=detail.policy_name,
        policy_description=detail.policy_description,
        policy_scope=detail.policy_scope,
        timeline=[TimelineIntervalOut(**asdict(t)) for t in detail.timeline],
        evidence=[EvidenceOut(**asdict(e)) for e in detail.evidence],
        is_weekend=detail.is_weekend,
        is_holiday=detail.is_holiday,
        holiday_name=detail.holiday_name,
        leave_name=detail.leave_name,
    )


@router.get("/export")
def export_calendar(
    user: Annotated[CurrentUser, Depends(current_user)],
    month: Annotated[str, Query(...)],
    employee_id: Annotated[Optional[int], Query()] = None,
    the_date: Annotated[
        Optional[date_type],
        Query(alias="date", description="Single-day export when present."),
    ] = None,
) -> StreamingResponse:
    """Excel export. Three modes:

    * ``employee_id`` + ``date``: single-day row for one employee
      (drawer's Export button).
    * ``employee_id`` only: per-person month view.
    * Neither: company-aggregate month view (Admin/HR/Manager).
    """

    try:
        month_start, month_end = queries.parse_month(month)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    scope = TenantScope(tenant_id=user.tenant_id)

    # Lazy openpyxl import keeps the request path cheap on cold boots.
    from openpyxl import Workbook  # noqa: PLC0415
    from openpyxl.styles import Font  # noqa: PLC0415

    wb = Workbook(write_only=False)
    ws = wb.active
    bold = Font(bold=True)

    if employee_id is not None:
        _check_can_view_employee(user, scope, employee_id)

    if employee_id is not None and the_date is not None:
        # Single-day mode (drawer Export button).
        with get_engine().begin() as conn:
            detail = queries.fetch_day_detail(
                conn, scope, employee_id=employee_id, the_date=the_date
            )
        if detail is None:
            raise HTTPException(status_code=404, detail="not found")
        ws.title = the_date.isoformat()
        for col, h in enumerate(
            ["Field", "Value"], start=1
        ):
            c = ws.cell(row=1, column=col, value=h)
            c.font = bold
        rows = [
            ("Employee", detail.full_name),
            ("Code", detail.employee_code),
            ("Department", detail.department_name),
            ("Date", detail.date.isoformat()),
            ("Status", detail.status),
            ("In time", detail.in_time or "—"),
            ("Out time", detail.out_time or "—"),
            (
                "Total minutes",
                detail.total_minutes if detail.total_minutes is not None else 0,
            ),
            ("Overtime minutes", detail.overtime_minutes),
            ("Policy", detail.policy_name or "—"),
            ("Holiday", detail.holiday_name or ""),
            ("Leave", detail.leave_name or ""),
            ("Evidence count", len(detail.evidence)),
        ]
        for i, (k, v) in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=k)
            ws.cell(row=i, column=2, value=v)
        filename = (
            f"hadir-attendance-{detail.employee_code}-"
            f"{detail.date.isoformat()}.xlsx"
        )
    elif employee_id is not None:
        # Per-person month export.
        with get_engine().begin() as conn:
            settings = load_tenant_settings(conn, scope)
            today_local = datetime.now(timezone.utc).astimezone(
                local_tz_for(settings)
            ).date()
            days = queries.person_view(
                conn,
                scope,
                employee_id=employee_id,
                month_start=month_start,
                month_end=month_end,
                today_local=today_local,
            )
            emp_row = conn.execute(
                select(
                    employees.c.employee_code, employees.c.full_name
                ).where(
                    employees.c.tenant_id == scope.tenant_id,
                    employees.c.id == employee_id,
                )
            ).first()
        ws.title = month
        for col, h in enumerate(
            [
                "Date", "Status", "In", "Out", "Total min",
                "Overtime min", "Policy", "Leave", "Holiday",
            ],
            start=1,
        ):
            c = ws.cell(row=1, column=col, value=h)
            c.font = bold
        for i, d in enumerate(days, start=2):
            ws.cell(row=i, column=1, value=d.date.isoformat())
            ws.cell(row=i, column=2, value=d.status)
            ws.cell(row=i, column=3, value=d.in_time or "")
            ws.cell(row=i, column=4, value=d.out_time or "")
            ws.cell(row=i, column=5, value=d.total_minutes or 0)
            ws.cell(row=i, column=6, value=d.overtime_minutes or 0)
            ws.cell(row=i, column=7, value=d.policy_name or "")
            ws.cell(row=i, column=8, value=d.leave_name or "")
            ws.cell(row=i, column=9, value=d.holiday_name or "")
        code = str(emp_row.employee_code) if emp_row else "unknown"
        filename = f"hadir-attendance-{code}-{month}.xlsx"
    else:
        # Company aggregate month export.
        if not (_is_admin_like(user) or _is_manager(user)):
            raise HTTPException(status_code=403, detail="forbidden")
        employee_ids = _resolve_visible_employees(user, scope)
        with get_engine().begin() as conn:
            days = queries.company_view(
                conn,
                scope,
                month_start=month_start,
                month_end=month_end,
                employee_ids=employee_ids,
            )
        ws.title = month
        for col, h in enumerate(
            [
                "Date", "Present", "Late", "Absent", "Leave",
                "Active employees", "% present", "Holiday",
                "Weekend",
            ],
            start=1,
        ):
            c = ws.cell(row=1, column=col, value=h)
            c.font = bold
        for i, d in enumerate(days, start=2):
            ws.cell(row=i, column=1, value=d.date.isoformat())
            ws.cell(row=i, column=2, value=d.present_count)
            ws.cell(row=i, column=3, value=d.late_count)
            ws.cell(row=i, column=4, value=d.absent_count)
            ws.cell(row=i, column=5, value=d.leave_count)
            ws.cell(row=i, column=6, value=d.active_employees)
            ws.cell(row=i, column=7, value=d.percent_present)
            ws.cell(row=i, column=8, value=d.holiday_name or "")
            ws.cell(row=i, column=9, value="weekend" if d.is_weekend else "")
        filename = f"hadir-attendance-company-{month}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Audit the export so an auditor can see who pulled what.
    with get_engine().begin() as conn:
        write_audit(
            conn,
            tenant_id=scope.tenant_id,
            actor_user_id=user.id if user.id > 0 else None,
            action="attendance_calendar.exported",
            entity_type="report",
            entity_id=None,
            after={
                "month": month,
                "employee_id": employee_id,
                "date": the_date.isoformat() if the_date else None,
            },
        )

    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )

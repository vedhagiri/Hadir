"""Excel import + export using openpyxl.

Kept tiny and dependency-light: no pandas, no abstraction framework. The
import parser normalises headers (``lower``, ``strip``, space→underscore)
so operator-friendly variations like "Employee Code" still match the
canonical ``employee_code`` column.

Row numbers reported in errors are **Excel row numbers** (1-indexed,
header is row 1, first data row is row 2). That's what a human opening
the file in Excel will see.

P12 wires custom fields into both directions: the export appends one
column per defined field (using its ``code`` as the header), and the
import accepts those same columns to populate ``custom_field_values``.
Unknown extra columns produce **row warnings**, not row errors —
operator forgot to delete a leftover column, the rest of the file
should still import.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date as _date
from io import BytesIO
from typing import Iterator, Optional

from openpyxl import Workbook, load_workbook

from maugood.employees.repository import EmployeeRow

REQUIRED_COLUMNS: tuple[str, ...] = (
    "employee_code",
    "full_name",
    "department_code",
)

# ``email`` used to be required. It's now optional — many customers
# import an HR roster that doesn't carry an email column at all
# (employee_code identifies the row), and the user-account creation
# happens later when an Admin promotes the employee to Manager / HR /
# Admin. The parser still picks ``email`` up when present and threads
# it through to ``ImportRow.email``.
_EMAIL_COLUMN: str = "email"

# P28.7 added six optional columns. They land in the export and the
# import accepts them; ``reports_to_email`` is resolved to a user_id at
# import time. ``status`` was already exported but not parsed —
# we still don't import status (operators set it via the API).
P28_7_OPTIONAL_COLUMNS: tuple[str, ...] = (
    "designation",
    "phone",
    "reports_to_email",
    "joining_date",
    "relieving_date",
    "status",
)

EXPORT_COLUMNS: tuple[str, ...] = (
    "employee_code",
    "full_name",
    "email",
    "department_code",
    "status",
    "photo_count",
    "designation",
    "phone",
    "reports_to_email",
    "joining_date",
    "relieving_date",
    "deactivation_reason",
)

# Headers we recognise as part of the employee row itself — anything else
# is treated as a candidate custom-field code on import.
_KNOWN_HEADERS: frozenset[str] = frozenset(
    {
        *REQUIRED_COLUMNS,
        _EMAIL_COLUMN,
        "photo_count",
        "deactivation_reason",
        *P28_7_OPTIONAL_COLUMNS,
    }
)


class ImportParseError(Exception):
    """Raised when the file itself is invalid (missing headers, not XLSX)."""


@dataclass(frozen=True, slots=True)
class ImportRow:
    """One parsed data row. ``excel_row`` is 1-indexed from the top of the file."""

    excel_row: int
    employee_code: str
    full_name: str
    email: Optional[str]
    department_code: str
    # P28.7 optional columns. Strings as the parser sees them; the
    # router coerces (date parsing, manager lookup) and validates.
    designation: Optional[str] = None
    phone: Optional[str] = None
    reports_to_email: Optional[str] = None
    joining_date: Optional[str] = None
    relieving_date: Optional[str] = None
    status: Optional[str] = None
    # ``custom_values`` keys are the **raw** header strings as they
    # appeared in the spreadsheet (already lower-snake-cased). The
    # router decides which match a known custom-field code, which
    # produce warnings, and how to coerce per type.
    custom_values: dict[str, str] = field(default_factory=dict)


def _normalise_header(raw: object) -> str:
    if raw is None:
        return ""
    return str(raw).strip().lower().replace(" ", "_")


def _cell_str(value: object) -> str:
    """Coerce a cell value to a trimmed string; empty for None/whitespace."""

    if value is None:
        return ""
    text = str(value).strip()
    return text


def parse_import(stream: BytesIO) -> Iterator[ImportRow]:
    """Yield ``ImportRow`` for each non-empty data row.

    Raises ``ImportParseError`` if the workbook can't be opened or if any
    required column is missing.
    """

    try:
        # read_only so we don't load the whole grid into memory for a
        # thousand-row file — openpyxl streams rows from the archive.
        wb = load_workbook(stream, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — openpyxl raises a zoo of types
        raise ImportParseError(f"could not read workbook: {exc}") from exc

    try:
        ws = wb.active
        if ws is None:
            raise ImportParseError("workbook has no active sheet")

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            raise ImportParseError("workbook is empty")

        headers = [_normalise_header(c) for c in header_row]
        missing = [c for c in REQUIRED_COLUMNS if c not in headers]
        if missing:
            raise ImportParseError(
                f"missing required column(s): {', '.join(missing)}"
            )

        idx = {name: headers.index(name) for name in REQUIRED_COLUMNS}
        # ``email`` and the P28.7 fields are all optional — locate them
        # when present, leave the slot None when the column is absent.
        opt_idx: dict[str, Optional[int]] = {
            col: (headers.index(col) if col in headers else None)
            for col in (_EMAIL_COLUMN, *P28_7_OPTIONAL_COLUMNS)
        }
        # Index every non-required, non-blank column too — these are the
        # custom-field candidates.
        custom_idx: dict[str, int] = {}
        for pos, header in enumerate(headers):
            if not header:
                continue
            if header in _KNOWN_HEADERS:
                continue
            if header in custom_idx:
                # Duplicate header — keep the first; the parser doesn't
                # have a per-row error channel for this and the operator
                # will see the warning channel via the unknown-code path
                # if applicable.
                continue
            custom_idx[header] = pos

        def _opt_cell(row, key: str) -> Optional[str]:
            pos = opt_idx[key]
            if pos is None or pos >= len(row):
                return None
            value = row[pos]
            if value is None:
                return None
            if hasattr(value, "isoformat"):
                text = value.isoformat()
            else:
                text = _cell_str(value)
            return text or None

        for excel_row, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            # Skip fully-empty rows so trailing blank lines don't generate
            # a dozen "missing employee_code" errors.
            if all(v is None or _cell_str(v) == "" for v in row):
                continue

            code = _cell_str(row[idx["employee_code"]]) if idx["employee_code"] < len(row) else ""
            name = _cell_str(row[idx["full_name"]]) if idx["full_name"] < len(row) else ""
            email_raw = _opt_cell(row, _EMAIL_COLUMN) or ""
            dept_code = _cell_str(row[idx["department_code"]]) if idx["department_code"] < len(row) else ""

            cv: dict[str, str] = {}
            for header, pos in custom_idx.items():
                if pos >= len(row):
                    continue
                # Preserve the raw cell value via str/strip — date/number
                # cells come through as Python types and the router
                # coerces them via the field's declared type.
                value = row[pos]
                if value is None:
                    continue
                if hasattr(value, "isoformat"):
                    text = value.isoformat()
                else:
                    text = _cell_str(value)
                if text:
                    cv[header] = text

            yield ImportRow(
                excel_row=excel_row,
                employee_code=code,
                full_name=name,
                email=email_raw or None,
                department_code=dept_code,
                designation=_opt_cell(row, "designation"),
                phone=_opt_cell(row, "phone"),
                reports_to_email=_opt_cell(row, "reports_to_email"),
                joining_date=_opt_cell(row, "joining_date"),
                relieving_date=_opt_cell(row, "relieving_date"),
                status=_opt_cell(row, "status"),
                custom_values=cv,
            )
    finally:
        wb.close()


def parse_csv_import(data: bytes) -> Iterator[ImportRow]:
    """CSV variant of ``parse_import``. Same column contract: required
    headers ``employee_code, full_name, email, department_code``;
    optional headers from ``P28_7_OPTIONAL_COLUMNS``; any other header
    becomes a custom-field candidate. Department codes must match an
    existing department row — that validation lives in the router.

    Accepts UTF-8 with optional BOM. Headers are case-and-whitespace-
    normalised the same way as the XLSX parser. Per-row failures
    surface in the router's response just like the XLSX path.
    """

    import csv as _csv  # noqa: PLC0415

    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ImportParseError("file must be UTF-8 encoded") from exc

    from io import StringIO  # noqa: PLC0415

    reader = _csv.reader(StringIO(text))
    iterator = iter(reader)
    try:
        header_row = next(iterator)
    except StopIteration as exc:
        raise ImportParseError("CSV is empty") from exc

    headers = [_normalise_header(c) for c in header_row]
    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        raise ImportParseError(
            f"missing required column(s): {', '.join(missing)}"
        )

    idx = {name: headers.index(name) for name in REQUIRED_COLUMNS}
    opt_idx: dict[str, Optional[int]] = {
        col: (headers.index(col) if col in headers else None)
        for col in P28_7_OPTIONAL_COLUMNS
    }
    custom_idx: dict[str, int] = {}
    for pos, header in enumerate(headers):
        if not header or header in _KNOWN_HEADERS or header in custom_idx:
            continue
        custom_idx[header] = pos

    def _opt(row, key: str) -> Optional[str]:
        pos = opt_idx[key]
        if pos is None or pos >= len(row):
            return None
        v = row[pos].strip()
        return v or None

    for excel_row, row in enumerate(iterator, start=2):
        if all((cell or "").strip() == "" for cell in row):
            continue
        code = (
            row[idx["employee_code"]].strip()
            if idx["employee_code"] < len(row)
            else ""
        )
        name = (
            row[idx["full_name"]].strip()
            if idx["full_name"] < len(row)
            else ""
        )
        email_raw = (
            row[idx["email"]].strip() if idx["email"] < len(row) else ""
        )
        dept_code = (
            row[idx["department_code"]].strip()
            if idx["department_code"] < len(row)
            else ""
        )

        cv: dict[str, str] = {}
        for header, pos in custom_idx.items():
            if pos < len(row):
                v = row[pos].strip()
                if v:
                    cv[header] = v

        yield ImportRow(
            excel_row=excel_row,
            employee_code=code,
            full_name=name,
            email=email_raw or None,
            department_code=dept_code,
            designation=_opt(row, "designation"),
            phone=_opt(row, "phone"),
            reports_to_email=_opt(row, "reports_to_email"),
            joining_date=_opt(row, "joining_date"),
            relieving_date=_opt(row, "relieving_date"),
            status=_opt(row, "status"),
            custom_values=cv,
        )


def build_export(
    rows: list[EmployeeRow],
    *,
    custom_field_codes: tuple[str, ...] = (),
    values_by_employee: Optional[dict[int, dict[str, str]]] = None,
    reports_to_email_by_user: Optional[dict[int, str]] = None,
) -> BytesIO:
    """Produce an XLSX in-memory.

    The base columns from the pilot stay first; ``custom_field_codes`` is
    appended in the order returned by the custom-fields repository
    (``display_order`` ascending). Cells lookup against
    ``values_by_employee[employee_id][code]`` — missing entries become
    empty cells.

    P28.7 adds the lifecycle + HR org-chart columns. ``reports_to_email``
    maps each ``reports_to_user_id`` to its email so re-importing the
    same XLSX round-trips cleanly.
    """

    values_by_employee = values_by_employee or {}
    reports_to_email_by_user = reports_to_email_by_user or {}

    wb = Workbook()
    ws = wb.active
    assert ws is not None  # openpyxl always returns an active sheet on a fresh WB
    ws.title = "Employees"

    headers = list(EXPORT_COLUMNS) + list(custom_field_codes)
    ws.append(headers)
    for row in rows:
        reports_to_email = (
            reports_to_email_by_user.get(row.reports_to_user_id, "")
            if row.reports_to_user_id is not None
            else ""
        )
        base = [
            row.employee_code,
            row.full_name,
            row.email or "",
            row.department_code,
            row.status,
            row.photo_count,
            row.designation or "",
            row.phone or "",
            reports_to_email,
            row.joining_date.isoformat() if row.joining_date else "",
            row.relieving_date.isoformat() if row.relieving_date else "",
            row.deactivation_reason or "",
        ]
        custom_cells = [
            values_by_employee.get(row.id, {}).get(code, "")
            for code in custom_field_codes
        ]
        ws.append(base + custom_cells)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_iso_date(value: Optional[str]) -> Optional[_date]:
    """Parse a date cell value (already stringified by ``parse_import``).

    Accepts ISO 8601 (``YYYY-MM-DD``) and the openpyxl-stringified
    full-datetime form (``YYYY-MM-DDTHH:MM:SS``). Returns None on empty
    input. Raises ``ValueError`` on a malformed value so the import
    handler can convert it to a per-row error.
    """

    if value is None or not str(value).strip():
        return None
    s = str(value).strip()
    if "T" in s:
        s = s.split("T", 1)[0]
    return _date.fromisoformat(s)

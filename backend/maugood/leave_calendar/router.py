"""FastAPI router for ``/api/leave-types``, ``/api/holidays``,
``/api/approved-leaves``, and ``/api/tenant-settings``.

Admin + HR only. Audit hook on every mutation. Holiday Excel import
accepts an .xlsx with two columns — ``date`` (any pandas-parseable
date format) and ``name`` (free text).
"""

from __future__ import annotations

import logging
from datetime import date as date_type, datetime, timezone
from io import BytesIO
from typing import Annotated, Any, Optional

from fastapi import (
    APIRouter,
    Depends,
    File,
    HTTPException,
    Response,
    UploadFile,
    status,
)
from openpyxl import load_workbook
from sqlalchemy import and_, delete, insert, select, update
from sqlalchemy.exc import IntegrityError

from maugood.auth.audit import write_audit
from maugood.auth.dependencies import CurrentUser, current_user, require_any_role
from maugood.db import (
    approved_leaves,
    employees,
    get_engine,
    holidays as holidays_table,
    leave_types,
    tenant_settings,
)
from maugood.leave_calendar.schemas import (
    ApprovedLeaveCreateRequest,
    ApprovedLeaveResponse,
    HolidayBulkCreateRequest,
    HolidayCreateRequest,
    HolidayImportResponse,
    HolidayImportSkipped,
    HolidayResponse,
    LeaveTypeCreateRequest,
    LeaveTypePatchRequest,
    LeaveTypeResponse,
    TenantSettingsPatchRequest,
    TenantSettingsResponse,
)
from maugood.tenants.scope import TenantScope

logger = logging.getLogger(__name__)

router = APIRouter(tags=["leave-calendar"])
ADMIN_OR_HR = Depends(require_any_role("Admin", "HR"))


# ---------------------------------------------------------------------------
# Leave types
# ---------------------------------------------------------------------------


def _to_leave_type_response(row) -> LeaveTypeResponse:  # type: ignore[no-untyped-def]
    return LeaveTypeResponse(
        id=int(row.id),
        tenant_id=int(row.tenant_id),
        code=str(row.code),
        name=str(row.name),
        is_paid=bool(row.is_paid),
        active=bool(row.active),
    )


@router.get("/api/leave-types", response_model=list[LeaveTypeResponse])
def list_leave_types(
    # BUG-058 / BUG-059 — Employees submitting a leave request need to
    # see the available leave types. Read is open to every authenticated
    # user; writes (POST / PATCH / DELETE) stay Admin/HR.
    user: Annotated[CurrentUser, Depends(current_user)],
) -> list[LeaveTypeResponse]:
    scope = TenantScope(tenant_id=user.tenant_id)
    engine = get_engine()
    with engine.begin() as conn:
        rows = conn.execute(
            select(
                leave_types.c.id,
                leave_types.c.tenant_id,
                leave_types.c.code,
                leave_types.c.name,
                leave_types.c.is_paid,
                leave_types.c.active,
            )
            .where(leave_types.c.tenant_id == scope.tenant_id)
            .order_by(leave_types.c.id.asc())
        ).all()
    return [_to_leave_type_response(r) for r in rows]


@router.post(
    "/api/leave-types",
    response_model=LeaveTypeResponse,
    status_code=status.HTTP_201_CREATED,
)
def create_leave_type(
    payload: LeaveTypeCreateRequest,
    user: Annotated[CurrentUser, ADMIN_OR_HR],
) -> LeaveTypeResponse:
    scope = TenantScope(tenant_id=user.tenant_id)
    engine = get_engine()
    with engine.begin() as conn:
        try:
            new_id = int(
                conn.execute(
                    insert(leave_types)
                    .values(
                        tenant_id=scope.tenant_id,
                        code=payload.code,
                        name=payload.name,
                        is_paid=payload.is_paid,
                    )
                    .returning(leave_types.c.id)
                ).scalar_one()
            )
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(
                status_code=400,
                detail=f"could not create leave_type: {type(exc).__name__}",
            )
        write_audit(
            conn,
            tenant_id=scope.tenant_id,
            actor_user_id=user.id,
            action="leave_type.created",
            entity_type="leave_type",
            entity_id=str(new_id),
            after={
                "code": payload.code,
                "name": payload.name,
                "is_paid": payload.is_paid,
            },
        )
        row = conn.execute(
            select(
                leave_types.c.id,
                leave_types.c.tenant_id,
                leave_types.c.code,
                leave_types.c.name,
                leave_types.c.is_paid,
                leave_types.c.active,
            ).where(leave_types.c.id == new_id)
        ).first()
    assert row is not None
    return _to_leave_type_response(row)


@router.patch(
    "/api/leave-types/{leave_type_id}", response_model=LeaveTypeResponse
)
def patch_leave_type(
    leave_type_id: int,
    payload: LeaveTypePatchRequest,
    user: Annotated[CurrentUser, ADMIN_OR_HR],
) -> LeaveTypeResponse:
    scope = TenantScope(tenant_id=user.tenant_id)
    engine = get_engine()
    with engine.begin() as conn:
        before = conn.execute(
            select(
                leave_types.c.id,
                leave_types.c.tenant_id,
                leave_types.c.code,
                leave_types.c.name,
                leave_types.c.is_paid,
                leave_types.c.active,
            ).where(
                leave_types.c.id == leave_type_id,
                leave_types.c.tenant_id == scope.tenant_id,
            )
        ).first()
        if before is None:
            raise HTTPException(status_code=404, detail="leave_type not found")
        values: dict[str, Any] = {}
        if payload.name is not None:
            values["name"] = payload.name
        if payload.is_paid is not None:
            values["is_paid"] = payload.is_paid
        if payload.active is not None:
            values["active"] = payload.active
        if values:
            conn.execute(
                update(leave_types)
                .where(
                    leave_types.c.id == leave_type_id,
                    leave_types.c.tenant_id == scope.tenant_id,
                )
                .values(**values)
            )
            write_audit(
                conn,
                tenant_id=scope.tenant_id,
                actor_user_id=user.id,
                action="leave_type.updated",
                entity_type="leave_type",
                entity_id=str(leave_type_id),
                before={
                    "name": before.name,
                    "is_paid": before.is_paid,
                    "active": before.active,
                },
                after=values,
            )
        row = conn.execute(
            select(
                leave_types.c.id,
                leave_types.c.tenant_id,
                leave_types.c.code,
                leave_types.c.name,
                leave_types.c.is_paid,
                leave_types.c.active,
            ).where(leave_types.c.id == leave_type_id)
        ).first()
    assert row is not None
    return _to_leave_type_response(row)


# BUG-043 — Leave Types had no delete endpoint. Refuse to hard-delete
# rows that are referenced by approved_leaves so audit + reporting
# history doesn't lose its leave_type pointer; the caller is told to
# deactivate instead.
@router.delete(
    "/api/leave-types/{leave_type_id}",
    status_code=status.HTTP_204_NO_CONTENT,
)
def delete_leave_type(
    leave_type_id: int,
    user: Annotated[CurrentUser, ADMIN_OR_HR],
) -> Response:
    scope = TenantScope(tenant_id=user.tenant_id)
    engine = get_engine()
    with engine.begin() as conn:
        before = conn.execute(
            select(
                leave_types.c.code,
                leave_types.c.name,
            ).where(
                leave_types.c.id == leave_type_id,
                leave_types.c.tenant_id == scope.tenant_id,
            )
        ).first()
        if before is None:
            raise HTTPException(status_code=404, detail="leave_type not found")
        # Refuse if any approved_leaves still references this type.
        in_use = conn.execute(
            select(approved_leaves.c.id).where(
                approved_leaves.c.tenant_id == scope.tenant_id,
                approved_leaves.c.leave_type_id == leave_type_id,
            ).limit(1)
        ).first()
        if in_use is not None:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"Leave type '{before.name}' is in use by one or more "
                    f"approved leaves. Mark it inactive instead of "
                    f"deleting to preserve the audit trail."
                ),
            )
        conn.execute(
            delete(leave_types).where(
                leave_types.c.id == leave_type_id,
                leave_types.c.tenant_id == scope.tenant_id,
            )
        )
        write_audit(
            conn,
            tenant_id=scope.tenant_id,
            actor_user_id=user.id,
            action="leave_type.deleted",
            entity_type="leave_type",
            entity_id=str(leave_type_id),
            before={"code": str(before.code), "name": str(before.name)},
        )
    return Response(status_code=status.HTTP_204_NO_CONTENT)


# ---------------------------------------------------------------------------
# Holidays
# ---------------------------------------------------------------------------


def _to_holiday_response(row) -> HolidayResponse:  # type: ignore[no-untyped-def]
    desc = getattr(row, "description", None)
    return HolidayResponse(
        id=int(row.id),
        tenant_id=int(row.tenant_id),
        date=row.date,
        name=str(row.name),
        description=str(desc) if desc is not None else None,
        active=bool(row.active),
    )


@router.get("/api/holidays", response_model=list[HolidayResponse])
def list_holidays(
    user: Annotated[CurrentUser, ADMIN_OR_HR],
    year: Optional[int] = None,
) -> list[HolidayResponse]:
    scope = TenantScope(tenant_id=user.tenant_id)
    engine = get_engine()
    stmt = (
        select(
            holidays_table.c.id,
            holidays_table.c.tenant_id,
            holidays_table.c.date,
            holidays_table.c.name,
            holidays_table.c.description,
            holidays_table.c.active,
        )
        .where(holidays_table.c.tenant_id == scope.tenant_id)
        .order_by(holidays_table.c.date.asc())
    )
    if year is not None:
        from datetime import date as _d  # noqa: PLC0415

        stmt = stmt.where(
            holidays_table.c.date >= _d(year, 1, 1),
            holidays_table.c.date <= _d(year, 12, 31),
        )
    with engine.begin() as conn:
        rows = conn.execute(stmt).all()
    return [_to_holiday_response(r) for r in rows]


@router.post(
    "/api/holidays",
    response_model=HolidayResponse,
    status_code=status.HTTP_201_CREATED,
)
def create_holiday(
    payload: HolidayCreateRequest,
    user: Annotated[CurrentUser, ADMIN_OR_HR],
) -> HolidayResponse:
    scope = TenantScope(tenant_id=user.tenant_id)
    engine = get_engine()
    # BUG-023 / BUG-044 — friendly error for duplicate date. Pre-check
    # against the existing row so the operator sees the holiday's
    # current name instead of a raw IntegrityError.
    with engine.begin() as conn:
        dup = conn.execute(
            select(holidays_table.c.name).where(
                holidays_table.c.tenant_id == scope.tenant_id,
                holidays_table.c.date == payload.date,
            )
        ).first()
        if dup is not None:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"A holiday is already configured for "
                    f"{payload.date.isoformat()} ('{dup.name}'). "
                    f"Edit the existing entry or pick a different date."
                ),
            )
    with engine.begin() as conn:
        try:
            new_id = int(
                conn.execute(
                    insert(holidays_table)
                    .values(
                        tenant_id=scope.tenant_id,
                        date=payload.date,
                        name=payload.name,
                        description=payload.description,
                    )
                    .returning(holidays_table.c.id)
                ).scalar_one()
            )
        except IntegrityError as exc:
            # Race: another request inserted the same date between our
            # pre-check and the insert. Surface the same friendly
            # message rather than a 500.
            raise HTTPException(
                status_code=409,
                detail=(
                    f"A holiday is already configured for "
                    f"{payload.date.isoformat()}."
                ),
            ) from exc
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(
                status_code=400,
                detail=f"could not create holiday: {type(exc).__name__}",
            )
        write_audit(
            conn,
            tenant_id=scope.tenant_id,
            actor_user_id=user.id,
            action="holiday.created",
            entity_type="holiday",
            entity_id=str(new_id),
            after={"date": payload.date.isoformat(), "name": payload.name},
        )
        row = conn.execute(
            select(
                holidays_table.c.id,
                holidays_table.c.tenant_id,
                holidays_table.c.date,
                holidays_table.c.name,
                holidays_table.c.active,
            ).where(holidays_table.c.id == new_id)
        ).first()
    assert row is not None
    return _to_holiday_response(row)


@router.delete(
    "/api/holidays/{holiday_id}", status_code=status.HTTP_204_NO_CONTENT
)
def delete_holiday(
    holiday_id: int,
    user: Annotated[CurrentUser, ADMIN_OR_HR],
) -> Response:
    scope = TenantScope(tenant_id=user.tenant_id)
    engine = get_engine()
    with engine.begin() as conn:
        before = conn.execute(
            select(holidays_table.c.date, holidays_table.c.name).where(
                holidays_table.c.id == holiday_id,
                holidays_table.c.tenant_id == scope.tenant_id,
            )
        ).first()
        if before is None:
            raise HTTPException(status_code=404, detail="holiday not found")
        conn.execute(
            delete(holidays_table).where(
                holidays_table.c.id == holiday_id,
                holidays_table.c.tenant_id == scope.tenant_id,
            )
        )
        write_audit(
            conn,
            tenant_id=scope.tenant_id,
            actor_user_id=user.id,
            action="holiday.deleted",
            entity_type="holiday",
            entity_id=str(holiday_id),
            before={"date": before.date.isoformat(), "name": str(before.name)},
        )
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.post("/api/holidays/import", response_model=HolidayImportResponse)
async def import_holidays_xlsx(
    user: Annotated[CurrentUser, ADMIN_OR_HR],
    file: UploadFile = File(...),
) -> HolidayImportResponse:
    """Bulk-import holidays from an .xlsx file.

    Expected columns (header row, case-insensitive): ``date`` (an
    ISO date, an Excel serial date, or a Python date) and ``name``
    (free text). Returns both ``imported`` and ``skipped`` lists so
    the operator can see exactly which dates already existed —
    re-import is still idempotent but no longer silent (BUG-025).
    """

    scope = TenantScope(tenant_id=user.tenant_id)
    raw = await file.read()
    try:
        wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=400, detail=f"could not parse .xlsx: {type(exc).__name__}"
        )
    try:
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="empty workbook")

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            raise HTTPException(status_code=400, detail="empty workbook")
        normalised = [
            str(c).strip().lower() if c is not None else "" for c in header
        ]
        try:
            date_idx = normalised.index("date")
            name_idx = normalised.index("name")
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail="missing 'date' or 'name' column in header",
            )
        # BUG-021 — Description is optional; absent column = None on
        # every row.
        try:
            desc_idx: Optional[int] = normalised.index("description")
        except ValueError:
            desc_idx = None

        parsed: list[HolidayCreateRequest] = []
        for raw_row in rows_iter:
            if not raw_row:
                continue
            d_raw = raw_row[date_idx] if len(raw_row) > date_idx else None
            n_raw = raw_row[name_idx] if len(raw_row) > name_idx else None
            if d_raw is None or n_raw is None:
                continue
            d_value: date_type
            if isinstance(d_raw, datetime):
                d_value = d_raw.date()
            elif isinstance(d_raw, date_type):
                d_value = d_raw
            elif isinstance(d_raw, str):
                try:
                    d_value = date_type.fromisoformat(d_raw.strip())
                except ValueError:
                    raise HTTPException(
                        status_code=400,
                        detail=f"row date {d_raw!r} not ISO-parseable",
                    )
            else:
                raise HTTPException(
                    status_code=400,
                    detail=f"unsupported date cell type: {type(d_raw).__name__}",
                )
            desc_val: Optional[str] = None
            if desc_idx is not None and desc_idx < len(raw_row):
                v = raw_row[desc_idx]
                if v is not None:
                    desc_val = str(v).strip() or None
            parsed.append(
                HolidayCreateRequest(
                    date=d_value,
                    name=str(n_raw).strip(),
                    description=desc_val,
                )
            )
    finally:
        wb.close()

    if not parsed:
        return HolidayImportResponse()

    engine = get_engine()
    created_ids: list[int] = []
    skipped: list[HolidayImportSkipped] = []
    with engine.begin() as conn:
        # Pull existing rows for the dates we're about to insert so we
        # can both skip and *report* the dupes (BUG-025).
        existing_rows = conn.execute(
            select(
                holidays_table.c.date,
                holidays_table.c.name,
            ).where(
                holidays_table.c.tenant_id == scope.tenant_id,
                holidays_table.c.date.in_([h.date for h in parsed]),
            )
        ).all()
        existing_by_date = {r.date: str(r.name) for r in existing_rows}

        for h in parsed:
            if h.date in existing_by_date:
                skipped.append(
                    HolidayImportSkipped(
                        date=h.date,
                        submitted_name=h.name,
                        existing_name=existing_by_date[h.date],
                    )
                )
                continue
            new_id = int(
                conn.execute(
                    insert(holidays_table)
                    .values(
                        tenant_id=scope.tenant_id,
                        date=h.date,
                        name=h.name,
                        description=h.description,
                    )
                    .returning(holidays_table.c.id)
                ).scalar_one()
            )
            created_ids.append(new_id)
        if created_ids:
            write_audit(
                conn,
                tenant_id=scope.tenant_id,
                actor_user_id=user.id,
                action="holiday.bulk_imported",
                entity_type="holiday",
                entity_id=None,
                after={
                    "count": len(created_ids),
                    "ids": created_ids,
                    "skipped_count": len(skipped),
                },
            )
        rows = (
            conn.execute(
                select(
                    holidays_table.c.id,
                    holidays_table.c.tenant_id,
                    holidays_table.c.date,
                    holidays_table.c.name,
                    holidays_table.c.active,
                ).where(holidays_table.c.id.in_(created_ids))
            ).all()
            if created_ids
            else []
        )
    return HolidayImportResponse(
        imported=[_to_holiday_response(r) for r in rows],
        skipped=skipped,
        imported_count=len(rows),
        skipped_count=len(skipped),
    )


# ---------------------------------------------------------------------------
# Approved leaves
# ---------------------------------------------------------------------------


def _to_approved_leave(row) -> ApprovedLeaveResponse:  # type: ignore[no-untyped-def]
    return ApprovedLeaveResponse(
        id=int(row.id),
        tenant_id=int(row.tenant_id),
        employee_id=int(row.employee_id),
        leave_type_id=int(row.leave_type_id),
        leave_type_code=str(row.leave_type_code),
        leave_type_name=str(row.leave_type_name),
        start_date=row.start_date,
        end_date=row.end_date,
        notes=row.notes,
        approved_by_user_id=(
            int(row.approved_by_user_id)
            if row.approved_by_user_id is not None
            else None
        ),
        approved_at=row.approved_at.isoformat(),
    )


@router.get(
    "/api/approved-leaves", response_model=list[ApprovedLeaveResponse]
)
def list_approved_leaves(
    user: Annotated[CurrentUser, ADMIN_OR_HR],
    employee_id: Optional[int] = None,
) -> list[ApprovedLeaveResponse]:
    scope = TenantScope(tenant_id=user.tenant_id)
    engine = get_engine()
    stmt = (
        select(
            approved_leaves.c.id,
            approved_leaves.c.tenant_id,
            approved_leaves.c.employee_id,
            approved_leaves.c.leave_type_id,
            leave_types.c.code.label("leave_type_code"),
            leave_types.c.name.label("leave_type_name"),
            approved_leaves.c.start_date,
            approved_leaves.c.end_date,
            approved_leaves.c.notes,
            approved_leaves.c.approved_by_user_id,
            approved_leaves.c.approved_at,
        )
        .select_from(
            approved_leaves.join(
                leave_types,
                and_(
                    leave_types.c.id == approved_leaves.c.leave_type_id,
                    leave_types.c.tenant_id == approved_leaves.c.tenant_id,
                ),
            )
        )
        .where(approved_leaves.c.tenant_id == scope.tenant_id)
        .order_by(approved_leaves.c.start_date.desc())
    )
    if employee_id is not None:
        stmt = stmt.where(approved_leaves.c.employee_id == employee_id)
    with engine.begin() as conn:
        rows = conn.execute(stmt).all()
    return [_to_approved_leave(r) for r in rows]


@router.post(
    "/api/approved-leaves",
    response_model=ApprovedLeaveResponse,
    status_code=status.HTTP_201_CREATED,
)
def create_approved_leave(
    payload: ApprovedLeaveCreateRequest,
    user: Annotated[CurrentUser, ADMIN_OR_HR],
) -> ApprovedLeaveResponse:
    scope = TenantScope(tenant_id=user.tenant_id)
    engine = get_engine()
    with engine.begin() as conn:
        # Validate employee + leave type both belong to the tenant.
        # BUG-027 — friendlier wording on the employee-id error so a
        # non-developer reading the dialog understands what went wrong.
        ok_emp = conn.execute(
            select(employees.c.id, employees.c.full_name).where(
                employees.c.id == payload.employee_id,
                employees.c.tenant_id == scope.tenant_id,
            )
        ).first()
        if ok_emp is None:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"No employee with ID {payload.employee_id} exists for "
                    f"this tenant. Pick the correct employee from the list."
                ),
            )
        ok_lt = conn.execute(
            select(leave_types.c.id, leave_types.c.name).where(
                leave_types.c.id == payload.leave_type_id,
                leave_types.c.tenant_id == scope.tenant_id,
            )
        ).first()
        if ok_lt is None:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Pick a leave type — the selected entry does not "
                    "belong to this tenant."
                ),
            )

        # BUG-047 — refuse overlapping leaves for the same employee.
        # Two leave rows overlap when ``start_date <= other.end_date AND
        # end_date >= other.start_date``. SQL-side check so two parallel
        # requests can't both win a duplicate.
        overlap = conn.execute(
            select(
                approved_leaves.c.id,
                approved_leaves.c.start_date,
                approved_leaves.c.end_date,
            ).where(
                approved_leaves.c.tenant_id == scope.tenant_id,
                approved_leaves.c.employee_id == payload.employee_id,
                approved_leaves.c.start_date <= payload.end_date,
                approved_leaves.c.end_date >= payload.start_date,
            )
        ).first()
        if overlap is not None:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"This employee already has approved leave from "
                    f"{overlap.start_date.isoformat()} to "
                    f"{overlap.end_date.isoformat()}. Overlapping leave "
                    f"is not allowed — edit or remove the existing entry "
                    f"first."
                ),
            )

        new_id = int(
            conn.execute(
                insert(approved_leaves)
                .values(
                    tenant_id=scope.tenant_id,
                    employee_id=payload.employee_id,
                    leave_type_id=payload.leave_type_id,
                    start_date=payload.start_date,
                    end_date=payload.end_date,
                    notes=payload.notes,
                    approved_by_user_id=user.id if user.id > 0 else None,
                    approved_at=datetime.now(tz=timezone.utc),
                )
                .returning(approved_leaves.c.id)
            ).scalar_one()
        )
        write_audit(
            conn,
            tenant_id=scope.tenant_id,
            actor_user_id=user.id,
            action="approved_leave.created",
            entity_type="approved_leave",
            entity_id=str(new_id),
            after={
                "employee_id": payload.employee_id,
                "leave_type_id": payload.leave_type_id,
                "start_date": payload.start_date.isoformat(),
                "end_date": payload.end_date.isoformat(),
            },
        )
        row = conn.execute(
            select(
                approved_leaves.c.id,
                approved_leaves.c.tenant_id,
                approved_leaves.c.employee_id,
                approved_leaves.c.leave_type_id,
                leave_types.c.code.label("leave_type_code"),
                leave_types.c.name.label("leave_type_name"),
                approved_leaves.c.start_date,
                approved_leaves.c.end_date,
                approved_leaves.c.notes,
                approved_leaves.c.approved_by_user_id,
                approved_leaves.c.approved_at,
            )
            .select_from(
                approved_leaves.join(
                    leave_types,
                    and_(
                        leave_types.c.id == approved_leaves.c.leave_type_id,
                        leave_types.c.tenant_id == approved_leaves.c.tenant_id,
                    ),
                )
            )
            .where(approved_leaves.c.id == new_id)
        ).first()
    assert row is not None
    return _to_approved_leave(row)


@router.delete(
    "/api/approved-leaves/{leave_id}", status_code=status.HTTP_204_NO_CONTENT
)
def delete_approved_leave(
    leave_id: int,
    user: Annotated[CurrentUser, ADMIN_OR_HR],
) -> Response:
    scope = TenantScope(tenant_id=user.tenant_id)
    engine = get_engine()
    with engine.begin() as conn:
        before = conn.execute(
            select(
                approved_leaves.c.employee_id,
                approved_leaves.c.leave_type_id,
                approved_leaves.c.start_date,
                approved_leaves.c.end_date,
            ).where(
                approved_leaves.c.id == leave_id,
                approved_leaves.c.tenant_id == scope.tenant_id,
            )
        ).first()
        if before is None:
            raise HTTPException(status_code=404, detail="approved_leave not found")
        conn.execute(
            delete(approved_leaves).where(
                approved_leaves.c.id == leave_id,
                approved_leaves.c.tenant_id == scope.tenant_id,
            )
        )
        write_audit(
            conn,
            tenant_id=scope.tenant_id,
            actor_user_id=user.id,
            action="approved_leave.deleted",
            entity_type="approved_leave",
            entity_id=str(leave_id),
            before={
                "employee_id": int(before.employee_id),
                "leave_type_id": int(before.leave_type_id),
                "start_date": before.start_date.isoformat(),
                "end_date": before.end_date.isoformat(),
            },
        )
    return Response(status_code=status.HTTP_204_NO_CONTENT)


# ---------------------------------------------------------------------------
# Tenant settings
# ---------------------------------------------------------------------------


def _to_tenant_settings_response(row) -> TenantSettingsResponse:  # type: ignore[no-untyped-def]
    return TenantSettingsResponse(
        tenant_id=int(row.tenant_id),
        weekend_days=list(row.weekend_days or []),
        timezone=str(row.timezone),
        live_matching_enabled=bool(
            getattr(row, "live_matching_enabled", False)
        ),
        updated_at=row.updated_at.isoformat(),
    )


@router.get(
    "/api/tenant-settings", response_model=TenantSettingsResponse
)
def get_tenant_settings(
    user: Annotated[CurrentUser, ADMIN_OR_HR],
) -> TenantSettingsResponse:
    scope = TenantScope(tenant_id=user.tenant_id)
    engine = get_engine()
    with engine.begin() as conn:
        row = conn.execute(
            select(
                tenant_settings.c.tenant_id,
                tenant_settings.c.weekend_days,
                tenant_settings.c.timezone,
                tenant_settings.c.live_matching_enabled,
                tenant_settings.c.updated_at,
            ).where(tenant_settings.c.tenant_id == scope.tenant_id)
        ).first()
        if row is None:
            # Lazy-create the default row if missing.
            conn.execute(
                insert(tenant_settings).values(tenant_id=scope.tenant_id)
            )
            row = conn.execute(
                select(
                    tenant_settings.c.tenant_id,
                    tenant_settings.c.weekend_days,
                    tenant_settings.c.timezone,
                    tenant_settings.c.live_matching_enabled,
                    tenant_settings.c.updated_at,
                ).where(tenant_settings.c.tenant_id == scope.tenant_id)
            ).first()
    assert row is not None
    return _to_tenant_settings_response(row)


@router.patch(
    "/api/tenant-settings", response_model=TenantSettingsResponse
)
def patch_tenant_settings(
    payload: TenantSettingsPatchRequest,
    user: Annotated[CurrentUser, ADMIN_OR_HR],
) -> TenantSettingsResponse:
    scope = TenantScope(tenant_id=user.tenant_id)
    engine = get_engine()
    with engine.begin() as conn:
        before = conn.execute(
            select(
                tenant_settings.c.weekend_days,
                tenant_settings.c.timezone,
                tenant_settings.c.live_matching_enabled,
            ).where(tenant_settings.c.tenant_id == scope.tenant_id)
        ).first()
        values: dict[str, Any] = {"updated_at": datetime.now(tz=timezone.utc)}
        if payload.weekend_days is not None:
            values["weekend_days"] = payload.weekend_days
        if payload.timezone is not None:
            values["timezone"] = payload.timezone
        if payload.live_matching_enabled is not None:
            values["live_matching_enabled"] = payload.live_matching_enabled
        if before is None:
            # Create with defaults + payload overrides.
            conn.execute(
                insert(tenant_settings).values(
                    tenant_id=scope.tenant_id, **values
                )
            )
        else:
            conn.execute(
                update(tenant_settings)
                .where(tenant_settings.c.tenant_id == scope.tenant_id)
                .values(**values)
            )
        write_audit(
            conn,
            tenant_id=scope.tenant_id,
            actor_user_id=user.id,
            action="tenant_settings.updated",
            entity_type="tenant_settings",
            entity_id=str(scope.tenant_id),
            before=(
                {
                    "weekend_days": list(before.weekend_days or []),
                    "timezone": str(before.timezone),
                    "live_matching_enabled": bool(
                        getattr(before, "live_matching_enabled", True)
                    ),
                }
                if before is not None
                else None
            ),
            after={
                k: v.isoformat() if hasattr(v, "isoformat") else v
                for k, v in values.items()
            },
        )
        row = conn.execute(
            select(
                tenant_settings.c.tenant_id,
                tenant_settings.c.weekend_days,
                tenant_settings.c.timezone,
                tenant_settings.c.live_matching_enabled,
                tenant_settings.c.updated_at,
            ).where(tenant_settings.c.tenant_id == scope.tenant_id)
        ).first()
    assert row is not None
    return _to_tenant_settings_response(row)

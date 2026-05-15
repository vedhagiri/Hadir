"""FastAPI router for ``/api/policies`` + ``/api/policy-assignments``.

Admin + HR can manage policies and assignments. Soft-delete by
setting ``active_until = today - 1`` (preserves history; the
attendance-records FK on ``shift_policies`` rejects a hard DELETE
anyway).

Audit:
* ``shift_policy.{created,updated,soft_deleted}``
* ``policy_assignment.{created,deleted}``
"""

from __future__ import annotations

import logging
from datetime import date as date_type, timedelta
from io import BytesIO
from typing import Annotated, Optional

from fastapi import (
    APIRouter,
    Depends,
    File,
    HTTPException,
    Response,
    UploadFile,
    status,
)
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy import and_, delete, func, insert, select, update

from maugood.auth.audit import write_audit
from maugood.auth.dependencies import CurrentUser, require_any_role
from maugood.db import (
    departments,
    employees,
    get_engine,
    policy_assignments,
    shift_policies,
)
from maugood.policies.schemas import (
    AssignmentCreateRequest,
    AssignmentResponse,
    PolicyCreateRequest,
    PolicyPatchRequest,
    PolicyResponse,
)
from maugood.tenants.scope import TenantScope

logger = logging.getLogger(__name__)

router = APIRouter(tags=["policies"])

# Admin + HR are both allowed; Manager / Employee are not.
ADMIN_OR_HR = Depends(require_any_role("Admin", "HR"))


# ---------------------------------------------------------------------------
# Policies
# ---------------------------------------------------------------------------


def _policy_to_response(row) -> PolicyResponse:  # type: ignore[no-untyped-def]
    return PolicyResponse(
        id=int(row.id),
        tenant_id=int(row.tenant_id),
        name=str(row.name),
        type=str(row.type),  # type: ignore[arg-type]
        config=dict(row.config or {}),
        active_from=row.active_from,
        active_until=row.active_until,
    )


@router.get("/api/policies", response_model=list[PolicyResponse])
def list_policies(user: Annotated[CurrentUser, ADMIN_OR_HR]) -> list[PolicyResponse]:
    scope = TenantScope(tenant_id=user.tenant_id)
    engine = get_engine()
    with engine.begin() as conn:
        rows = conn.execute(
            select(
                shift_policies.c.id,
                shift_policies.c.tenant_id,
                shift_policies.c.name,
                shift_policies.c.type,
                shift_policies.c.config,
                shift_policies.c.active_from,
                shift_policies.c.active_until,
            )
            .where(shift_policies.c.tenant_id == scope.tenant_id)
            .order_by(shift_policies.c.id.asc())
        ).all()
    return [_policy_to_response(r) for r in rows]


@router.post(
    "/api/policies",
    response_model=PolicyResponse,
    status_code=status.HTTP_201_CREATED,
)
def create_policy(
    payload: PolicyCreateRequest,
    user: Annotated[CurrentUser, ADMIN_OR_HR],
) -> PolicyResponse:
    scope = TenantScope(tenant_id=user.tenant_id)
    engine = get_engine()
    with engine.begin() as conn:
        # BUG-042 — duplicate-policy-name guard. Compare case-
        # insensitively against active (non-soft-deleted) rows so the
        # operator sees a friendly 409 instead of accidentally creating
        # two "Default 07:30–15:30" policies for the same tenant.
        existing = conn.execute(
            select(shift_policies.c.id, shift_policies.c.name).where(
                shift_policies.c.tenant_id == scope.tenant_id,
                func.lower(shift_policies.c.name) == payload.name.strip().lower(),
            )
        ).first()
        if existing is not None:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"A shift policy named '{existing.name}' already exists "
                    f"for this tenant. Pick a different name or edit the "
                    f"existing entry."
                ),
            )
        new_id = int(
            conn.execute(
                insert(shift_policies)
                .values(
                    tenant_id=scope.tenant_id,
                    name=payload.name,
                    type=payload.type,
                    config=payload.config.model_dump(exclude_none=True),
                    active_from=payload.active_from,
                    active_until=payload.active_until,
                )
                .returning(shift_policies.c.id)
            ).scalar_one()
        )
        write_audit(
            conn,
            tenant_id=scope.tenant_id,
            actor_user_id=user.id,
            action="shift_policy.created",
            entity_type="shift_policy",
            entity_id=str(new_id),
            after={
                "name": payload.name,
                "type": payload.type,
                "active_from": payload.active_from.isoformat(),
                "active_until": (
                    payload.active_until.isoformat()
                    if payload.active_until is not None
                    else None
                ),
            },
        )
        row = conn.execute(
            select(
                shift_policies.c.id,
                shift_policies.c.tenant_id,
                shift_policies.c.name,
                shift_policies.c.type,
                shift_policies.c.config,
                shift_policies.c.active_from,
                shift_policies.c.active_until,
            ).where(shift_policies.c.id == new_id)
        ).first()
    assert row is not None
    return _policy_to_response(row)


@router.patch("/api/policies/{policy_id}", response_model=PolicyResponse)
def patch_policy(
    policy_id: int,
    payload: PolicyPatchRequest,
    user: Annotated[CurrentUser, ADMIN_OR_HR],
) -> PolicyResponse:
    scope = TenantScope(tenant_id=user.tenant_id)
    engine = get_engine()
    with engine.begin() as conn:
        before = conn.execute(
            select(
                shift_policies.c.id,
                shift_policies.c.tenant_id,
                shift_policies.c.name,
                shift_policies.c.type,
                shift_policies.c.config,
                shift_policies.c.active_from,
                shift_policies.c.active_until,
            ).where(
                shift_policies.c.id == policy_id,
                shift_policies.c.tenant_id == scope.tenant_id,
            )
        ).first()
        if before is None:
            raise HTTPException(status_code=404, detail="policy not found")

        values: dict[str, object] = {}
        if payload.name is not None:
            values["name"] = payload.name
        if payload.config is not None:
            values["config"] = payload.config.model_dump(exclude_none=True)
        if payload.active_from is not None:
            values["active_from"] = payload.active_from
        if "active_until" in payload.model_fields_set:
            values["active_until"] = payload.active_until

        if values:
            conn.execute(
                update(shift_policies)
                .where(
                    shift_policies.c.id == policy_id,
                    shift_policies.c.tenant_id == scope.tenant_id,
                )
                .values(**values)
            )
            write_audit(
                conn,
                tenant_id=scope.tenant_id,
                actor_user_id=user.id,
                action="shift_policy.updated",
                entity_type="shift_policy",
                entity_id=str(policy_id),
                before={
                    "name": str(before.name),
                    "active_from": before.active_from.isoformat(),
                    "active_until": (
                        before.active_until.isoformat()
                        if before.active_until is not None
                        else None
                    ),
                },
                after={
                    k: (v.isoformat() if hasattr(v, "isoformat") else v)
                    for k, v in values.items()
                },
            )
        row = conn.execute(
            select(
                shift_policies.c.id,
                shift_policies.c.tenant_id,
                shift_policies.c.name,
                shift_policies.c.type,
                shift_policies.c.config,
                shift_policies.c.active_from,
                shift_policies.c.active_until,
            ).where(shift_policies.c.id == policy_id)
        ).first()
    assert row is not None
    return _policy_to_response(row)


@router.delete(
    "/api/policies/{policy_id}", status_code=status.HTTP_204_NO_CONTENT
)
def soft_delete_policy(
    policy_id: int,
    user: Annotated[CurrentUser, ADMIN_OR_HR],
) -> Response:
    """Soft-delete: sets ``active_until = today - 1`` so resolution skips it.

    Hard delete is refused — ``attendance_records.policy_id`` has a
    RESTRICT FK to ``shift_policies`` so historical rows always tie
    back to their original policy. Operators rely on that for audits.
    """

    scope = TenantScope(tenant_id=user.tenant_id)
    engine = get_engine()
    yesterday = date_type.today() - timedelta(days=1)
    with engine.begin() as conn:
        before = conn.execute(
            select(
                shift_policies.c.active_until,
                shift_policies.c.name,
            ).where(
                shift_policies.c.id == policy_id,
                shift_policies.c.tenant_id == scope.tenant_id,
            )
        ).first()
        if before is None:
            raise HTTPException(status_code=404, detail="policy not found")
        conn.execute(
            update(shift_policies)
            .where(
                shift_policies.c.id == policy_id,
                shift_policies.c.tenant_id == scope.tenant_id,
            )
            .values(active_until=yesterday)
        )
        write_audit(
            conn,
            tenant_id=scope.tenant_id,
            actor_user_id=user.id,
            action="shift_policy.soft_deleted",
            entity_type="shift_policy",
            entity_id=str(policy_id),
            before={
                "name": str(before.name),
                "active_until": (
                    before.active_until.isoformat()
                    if before.active_until is not None
                    else None
                ),
            },
            after={"active_until": yesterday.isoformat()},
        )
    return Response(status_code=status.HTTP_204_NO_CONTENT)


# ---------------------------------------------------------------------------
# BUG-040 — Shift policy XLSX import
# ---------------------------------------------------------------------------
#
# Expected columns (header row, case-insensitive):
#   name              required, ≤200 chars
#   type              one of Fixed / Flex / Ramadan / Custom (default Fixed)
#   start             HH:MM  (Fixed / Ramadan / Custom-Fixed)
#   end               HH:MM
#   grace_minutes     int    (Fixed) — defaults to 15
#   required_hours    int    — defaults to 8
#   active_from       YYYY-MM-DD (defaults to today)
#
# Rows whose name already exists for the tenant are reported as
# ``skipped`` instead of failing the whole batch (mirrors the holiday
# import shape).


class PolicyImportSkipped(BaseModel):
    row_number: int
    submitted_name: str
    reason: str


class PolicyImportResponse(BaseModel):
    imported: list[PolicyResponse] = []
    skipped: list[PolicyImportSkipped] = []
    imported_count: int = 0
    skipped_count: int = 0


class PolicyImportPreviewRow(BaseModel):
    """One importable row, validated. Surfaced in the preview UI."""

    row: int
    name: str
    type: str
    start: Optional[str] = None
    end: Optional[str] = None
    grace_minutes: Optional[int] = None
    required_hours: int
    active_from: str
    will_skip: bool = False
    skip_reason: Optional[str] = None


class PolicyImportPreviewError(BaseModel):
    row: int
    message: str


class PolicyImportPreviewResult(BaseModel):
    rows: list[PolicyImportPreviewRow] = []
    errors: list[PolicyImportPreviewError] = []


def _parse_policy_xlsx(
    raw: bytes,
) -> tuple[list[tuple[int, "PolicyCreateRequest"]], list[PolicyImportPreviewError]]:
    """Parse a Shift Policies XLSX into validated PolicyCreateRequest
    objects + per-row errors.

    The preview endpoint uses both arrays. The import endpoint uses the
    rows for INSERTs and surfaces errors in its response too. Shared
    so the two paths can't drift in behaviour.
    """

    parsed: list[tuple[int, "PolicyCreateRequest"]] = []
    errors: list[PolicyImportPreviewError] = []

    try:
        wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=400,
            detail=f"could not parse .xlsx: {type(exc).__name__}",
        )
    try:
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="empty workbook")
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            raise HTTPException(status_code=400, detail="empty workbook")
        norm = [
            str(c).strip().lower() if c is not None else "" for c in header
        ]

        def _idx(name: str) -> Optional[int]:
            try:
                return norm.index(name)
            except ValueError:
                return None

        i_name = _idx("name")
        if i_name is None:
            raise HTTPException(
                status_code=400, detail="missing 'name' column in header"
            )
        i_type = _idx("type")
        i_start = _idx("start")
        i_end = _idx("end")
        i_grace = _idx("grace_minutes")
        i_hours = _idx("required_hours")
        i_active_from = _idx("active_from")

        def _cell(row_tuple: tuple, idx: Optional[int]) -> Optional[str]:
            if idx is None or idx >= len(row_tuple):
                return None
            v = row_tuple[idx]
            if v is None:
                return None
            return str(v).strip()

        from datetime import date as _date_type  # noqa: PLC0415

        from maugood.policies.schemas import PolicyConfig  # noqa: PLC0415

        for row_number, raw_row in enumerate(rows_iter, start=2):
            if not raw_row:
                continue
            name = _cell(raw_row, i_name)
            if not name:
                # Blank ``name`` cell skips silently — typical when
                # operators leave trailing empty rows in the workbook.
                continue
            ptype_str = (_cell(raw_row, i_type) or "Fixed").capitalize()
            if ptype_str not in ("Fixed", "Flex", "Ramadan", "Custom"):
                errors.append(PolicyImportPreviewError(
                    row=row_number,
                    message=f"unsupported type {ptype_str!r} (valid: Fixed/Flex/Ramadan/Custom)",
                ))
                continue
            start = _cell(raw_row, i_start)
            end = _cell(raw_row, i_end)
            grace = _cell(raw_row, i_grace)
            hours = _cell(raw_row, i_hours)
            af_raw = _cell(raw_row, i_active_from)
            try:
                af = _date_type.fromisoformat(af_raw) if af_raw else _date_type.today()
            except ValueError:
                errors.append(PolicyImportPreviewError(
                    row=row_number,
                    message=f"active_from {af_raw!r} not ISO-parseable (YYYY-MM-DD)",
                ))
                continue

            cfg_kwargs: dict = {
                "required_hours": int(hours) if hours else 8,
            }
            if ptype_str in ("Fixed", "Ramadan"):
                if not (start and end):
                    errors.append(PolicyImportPreviewError(
                        row=row_number,
                        message=f"{ptype_str} requires start + end time (HH:MM)",
                    ))
                    continue
                cfg_kwargs["start"] = start
                cfg_kwargs["end"] = end
                cfg_kwargs["grace_minutes"] = int(grace) if grace else 15

            try:
                cfg = PolicyConfig(**cfg_kwargs)
                req = PolicyCreateRequest(
                    name=name,
                    type=ptype_str,
                    config=cfg,
                    active_from=af,
                )
            except Exception as exc:  # noqa: BLE001
                errors.append(PolicyImportPreviewError(
                    row=row_number, message=str(exc),
                ))
                continue
            parsed.append((row_number, req))
    finally:
        wb.close()

    return parsed, errors


@router.get("/api/policies/import-template")
def policies_import_template(
    user: Annotated[CurrentUser, ADMIN_OR_HR],
) -> StreamingResponse:
    """Stream a sample shift-policies workbook with one example row
    per type (Fixed / Flex / Ramadan / Custom) + a Field guide sheet.
    Operator clicks Download in the import modal, edits the file,
    uploads it back through the preview/import flow."""

    from openpyxl import Workbook  # noqa: PLC0415

    wb = Workbook()
    # Data sheet
    ws = wb.active
    ws.title = "Policies"
    headers = [
        "name", "type", "start", "end",
        "grace_minutes", "required_hours", "active_from",
    ]
    ws.append(headers)
    ws.append(["Standard Shift",   "Fixed",   "07:30", "15:30", 15, 8, "2026-01-01"])
    ws.append(["Flex Office",      "Flex",    "",      "",      "", 8, "2026-01-01"])
    ws.append(["Ramadan 2026",     "Ramadan", "09:00", "14:30", 10, 6, "2026-03-01"])
    ws.append(["Eid Custom",       "Custom",  "",      "",      "", 8, "2026-04-10"])

    # Field guide sheet — same UX as the employee import template.
    guide = wb.create_sheet("Field guide")
    guide.append(["Column", "Required?", "Notes"])
    guide.append([
        "name",
        "yes",
        "Unique within the tenant. Re-importing a row with an existing name is skipped (not failed).",
    ])
    guide.append([
        "type",
        "yes",
        "One of Fixed / Flex / Ramadan / Custom. Defaults to Fixed when blank.",
    ])
    guide.append([
        "start",
        "Fixed / Ramadan only",
        "Wall-clock start time HH:MM, 24-hour.",
    ])
    guide.append([
        "end",
        "Fixed / Ramadan only",
        "Wall-clock end time HH:MM, 24-hour.",
    ])
    guide.append([
        "grace_minutes",
        "optional",
        "Minutes of grace before late / after-end early-out. 0–180. Defaults to 15 for Fixed/Ramadan.",
    ])
    guide.append([
        "required_hours",
        "optional",
        "Daily required work hours. 1–24. Defaults to 8.",
    ])
    guide.append([
        "active_from",
        "optional",
        "ISO date YYYY-MM-DD. Defaults to today if blank.",
    ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                'attachment; filename="shift-policies-import-template.xlsx"'
            )
        },
    )


@router.post(
    "/api/policies/import-preview",
    response_model=PolicyImportPreviewResult,
)
async def import_policies_preview(
    user: Annotated[CurrentUser, ADMIN_OR_HR],
    file: UploadFile = File(...),
) -> PolicyImportPreviewResult:
    """Dry-run a shift-policies import: parse the workbook, validate
    each row, and report which rows would be skipped because their
    name already exists. **No DB writes.**

    The frontend posts the same file twice — once here for the
    preview, then to ``POST /api/policies/import`` after the
    operator confirms. Mirrors the employees import flow.
    """

    scope = TenantScope(tenant_id=user.tenant_id)
    raw = await file.read()
    parsed, errors = _parse_policy_xlsx(raw)

    # Snapshot existing names so the preview can flag "will be skipped".
    existing_names: set[str] = set()
    if parsed:
        engine = get_engine()
        with engine.begin() as conn:
            existing_names = {
                r.name.strip().lower()
                for r in conn.execute(
                    select(shift_policies.c.name).where(
                        shift_policies.c.tenant_id == scope.tenant_id,
                    )
                ).all()
            }

    rows: list[PolicyImportPreviewRow] = []
    for row_number, req in parsed:
        will_skip = req.name.strip().lower() in existing_names
        cfg = req.config
        rows.append(PolicyImportPreviewRow(
            row=row_number,
            name=req.name,
            type=req.type,
            start=cfg.start,
            end=cfg.end,
            grace_minutes=cfg.grace_minutes,
            required_hours=cfg.required_hours,
            active_from=req.active_from.isoformat(),
            will_skip=will_skip,
            skip_reason=(
                "A policy with this name already exists."
                if will_skip else None
            ),
        ))
    return PolicyImportPreviewResult(rows=rows, errors=errors)


@router.post("/api/policies/import", response_model=PolicyImportResponse)
async def import_policies_xlsx(
    user: Annotated[CurrentUser, ADMIN_OR_HR],
    file: UploadFile = File(...),
) -> PolicyImportResponse:
    """Bulk-create Fixed/Flex/Ramadan/Custom policies from an .xlsx
    file. Rows whose ``name`` already exists for the tenant are
    skipped (not failed); every parse error is collected so the
    operator can see exactly which rows didn't import (the modal
    surfaces them in the result panel)."""

    scope = TenantScope(tenant_id=user.tenant_id)
    raw = await file.read()
    parsed, parse_errors = _parse_policy_xlsx(raw)

    # Translate parse errors into skipped rows so the response shape
    # stays uniform — the operator sees one list, not two.
    skipped: list[PolicyImportSkipped] = [
        PolicyImportSkipped(
            row_number=e.row,
            submitted_name="(parse error)",
            reason=e.message,
        )
        for e in parse_errors
    ]

    if not parsed:
        return PolicyImportResponse(
            skipped=skipped, skipped_count=len(skipped),
        )

    engine = get_engine()
    imported: list[PolicyResponse] = []
    skipped: list[PolicyImportSkipped] = []
    with engine.begin() as conn:
        # Snapshot existing policy names (case-folded) so a re-import
        # is idempotent and surfaces per-row what was skipped.
        existing_names = {
            r.name.strip().lower()
            for r in conn.execute(
                select(shift_policies.c.name).where(
                    shift_policies.c.tenant_id == scope.tenant_id,
                )
            ).all()
        }
        for row_number, req in parsed:
            if req.name.strip().lower() in existing_names:
                skipped.append(
                    PolicyImportSkipped(
                        row_number=row_number,
                        submitted_name=req.name,
                        reason="A policy with this name already exists.",
                    )
                )
                continue
            new_id = int(
                conn.execute(
                    insert(shift_policies)
                    .values(
                        tenant_id=scope.tenant_id,
                        name=req.name,
                        type=req.type,
                        config=req.config.model_dump(exclude_none=True),
                        active_from=req.active_from,
                        active_until=req.active_until,
                    )
                    .returning(shift_policies.c.id)
                ).scalar_one()
            )
            existing_names.add(req.name.strip().lower())
            row = conn.execute(
                select(
                    shift_policies.c.id,
                    shift_policies.c.tenant_id,
                    shift_policies.c.name,
                    shift_policies.c.type,
                    shift_policies.c.config,
                    shift_policies.c.active_from,
                    shift_policies.c.active_until,
                ).where(shift_policies.c.id == new_id)
            ).first()
            assert row is not None
            imported.append(_policy_to_response(row))
        if imported:
            write_audit(
                conn,
                tenant_id=scope.tenant_id,
                actor_user_id=user.id,
                action="shift_policy.bulk_imported",
                entity_type="shift_policy",
                entity_id=None,
                after={
                    "count": len(imported),
                    "ids": [int(p.id) for p in imported],
                    "skipped_count": len(skipped),
                },
            )

    return PolicyImportResponse(
        imported=imported,
        skipped=skipped,
        imported_count=len(imported),
        skipped_count=len(skipped),
    )


# ---------------------------------------------------------------------------
# Policy assignments
# ---------------------------------------------------------------------------


def _assignment_to_response(row) -> AssignmentResponse:  # type: ignore[no-untyped-def]
    return AssignmentResponse(
        id=int(row.id),
        tenant_id=int(row.tenant_id),
        policy_id=int(row.policy_id),
        scope_type=str(row.scope_type),  # type: ignore[arg-type]
        scope_id=int(row.scope_id) if row.scope_id is not None else None,
        active_from=row.active_from,
        active_until=row.active_until,
    )


@router.get(
    "/api/policy-assignments", response_model=list[AssignmentResponse]
)
def list_assignments(
    user: Annotated[CurrentUser, ADMIN_OR_HR],
) -> list[AssignmentResponse]:
    scope = TenantScope(tenant_id=user.tenant_id)
    engine = get_engine()
    with engine.begin() as conn:
        rows = conn.execute(
            select(
                policy_assignments.c.id,
                policy_assignments.c.tenant_id,
                policy_assignments.c.policy_id,
                policy_assignments.c.scope_type,
                policy_assignments.c.scope_id,
                policy_assignments.c.active_from,
                policy_assignments.c.active_until,
            )
            .where(policy_assignments.c.tenant_id == scope.tenant_id)
            .order_by(
                policy_assignments.c.policy_id.asc(),
                policy_assignments.c.id.asc(),
            )
        ).all()
    return [_assignment_to_response(r) for r in rows]


def _validate_scope(
    conn,
    scope: TenantScope,
    *,
    scope_type: str,
    scope_id: int | None,
) -> None:
    if scope_type == "department":
        ok = conn.execute(
            select(departments.c.id).where(
                departments.c.id == scope_id,
                departments.c.tenant_id == scope.tenant_id,
            )
        ).first()
        if ok is None:
            raise HTTPException(
                status_code=400,
                detail="scope_id is not a department in this tenant",
            )
    elif scope_type == "employee":
        ok = conn.execute(
            select(employees.c.id).where(
                employees.c.id == scope_id,
                employees.c.tenant_id == scope.tenant_id,
            )
        ).first()
        if ok is None:
            raise HTTPException(
                status_code=400,
                detail="scope_id is not an employee in this tenant",
            )


@router.post(
    "/api/policy-assignments",
    response_model=AssignmentResponse,
    status_code=status.HTTP_201_CREATED,
)
def create_assignment(
    payload: AssignmentCreateRequest,
    user: Annotated[CurrentUser, ADMIN_OR_HR],
) -> AssignmentResponse:
    scope = TenantScope(tenant_id=user.tenant_id)
    engine = get_engine()
    with engine.begin() as conn:
        # Policy must exist in this tenant.
        ok = conn.execute(
            select(shift_policies.c.id).where(
                shift_policies.c.id == payload.policy_id,
                shift_policies.c.tenant_id == scope.tenant_id,
            )
        ).first()
        if ok is None:
            raise HTTPException(
                status_code=400,
                detail="policy_id is not a policy in this tenant",
            )
        _validate_scope(
            conn,
            scope,
            scope_type=payload.scope_type,
            scope_id=payload.scope_id,
        )
        new_id = int(
            conn.execute(
                insert(policy_assignments)
                .values(
                    tenant_id=scope.tenant_id,
                    policy_id=payload.policy_id,
                    scope_type=payload.scope_type,
                    scope_id=payload.scope_id,
                    active_from=payload.active_from,
                    active_until=payload.active_until,
                )
                .returning(policy_assignments.c.id)
            ).scalar_one()
        )
        write_audit(
            conn,
            tenant_id=scope.tenant_id,
            actor_user_id=user.id,
            action="policy_assignment.created",
            entity_type="policy_assignment",
            entity_id=str(new_id),
            after={
                "policy_id": payload.policy_id,
                "scope_type": payload.scope_type,
                "scope_id": payload.scope_id,
                "active_from": payload.active_from.isoformat(),
                "active_until": (
                    payload.active_until.isoformat()
                    if payload.active_until is not None
                    else None
                ),
            },
        )
        row = conn.execute(
            select(
                policy_assignments.c.id,
                policy_assignments.c.tenant_id,
                policy_assignments.c.policy_id,
                policy_assignments.c.scope_type,
                policy_assignments.c.scope_id,
                policy_assignments.c.active_from,
                policy_assignments.c.active_until,
            ).where(policy_assignments.c.id == new_id)
        ).first()
    assert row is not None
    return _assignment_to_response(row)


@router.delete(
    "/api/policy-assignments/{assignment_id}",
    status_code=status.HTTP_204_NO_CONTENT,
)
def delete_assignment(
    assignment_id: int,
    user: Annotated[CurrentUser, ADMIN_OR_HR],
) -> Response:
    scope = TenantScope(tenant_id=user.tenant_id)
    engine = get_engine()
    with engine.begin() as conn:
        before = conn.execute(
            select(
                policy_assignments.c.policy_id,
                policy_assignments.c.scope_type,
                policy_assignments.c.scope_id,
            ).where(
                policy_assignments.c.id == assignment_id,
                policy_assignments.c.tenant_id == scope.tenant_id,
            )
        ).first()
        if before is None:
            raise HTTPException(status_code=404, detail="assignment not found")
        conn.execute(
            delete(policy_assignments).where(
                policy_assignments.c.id == assignment_id,
                policy_assignments.c.tenant_id == scope.tenant_id,
            )
        )
        write_audit(
            conn,
            tenant_id=scope.tenant_id,
            actor_user_id=user.id,
            action="policy_assignment.deleted",
            entity_type="policy_assignment",
            entity_id=str(assignment_id),
            before={
                "policy_id": int(before.policy_id),
                "scope_type": str(before.scope_type),
                "scope_id": (
                    int(before.scope_id) if before.scope_id is not None else None
                ),
            },
        )
    return Response(status_code=status.HTTP_204_NO_CONTENT)

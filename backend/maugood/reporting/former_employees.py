"""P28.7 — "Former employees seen on premises" report.

Surfaces every ``detection_events`` row where ``former_employee_match=true``
joined to the matched (inactive) employee's snapshot — code, name,
deactivation reason, deactivation date.

Two output formats:

* ``json``  — list of rows, ready for the frontend table view.
* ``xlsx``  — in-memory spreadsheet for the operator's records.

HR + Admin only. The endpoint is mounted under ``/api/reports/`` to
share the same role gates + audit pattern as the existing reports.
"""

from __future__ import annotations

import logging
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from typing import Annotated, Literal, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import and_, select

from maugood.attendance.repository import load_tenant_settings, local_tz_for
from maugood.auth.audit import write_audit
from maugood.auth.dependencies import CurrentUser, require_any_role
from maugood.db import (
    cameras,
    detection_events,
    employees,
    get_engine,
)
from maugood.tenants.scope import TenantScope

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/reports", tags=["reports", "former-employees"])

ADMIN_OR_HR = Depends(require_any_role("Admin", "HR"))


class FormerEmployeeSighting(BaseModel):
    detection_event_id: int
    captured_at: datetime
    camera_id: Optional[int]
    camera_name: Optional[str]
    former_employee_id: Optional[int]
    former_employee_code: Optional[str]
    former_employee_name: Optional[str]
    confidence: Optional[float]
    deactivation_reason: Optional[str]
    deactivated_at: Optional[datetime]


class FormerEmployeeSightingsOut(BaseModel):
    items: list[FormerEmployeeSighting]
    total: int
    from_date: date
    to_date: date


def _query_sightings(
    scope: TenantScope, *, from_date: date, to_date: date
) -> list[FormerEmployeeSighting]:
    """Run the join + return ordered rows (newest first)."""

    engine = get_engine()
    with engine.begin() as conn:
        settings = load_tenant_settings(conn, scope)
        tz = local_tz_for(settings)

        # Convert local-day bounds to UTC for the captured_at comparison.
        start_utc = datetime.combine(
            from_date, datetime.min.time(), tzinfo=tz
        ).astimezone(timezone.utc)
        end_utc = datetime.combine(
            to_date + timedelta(days=1), datetime.min.time(), tzinfo=tz
        ).astimezone(timezone.utc)

        rows = conn.execute(
            select(
                detection_events.c.id,
                detection_events.c.captured_at,
                detection_events.c.camera_id,
                cameras.c.name.label("camera_name"),
                detection_events.c.former_match_employee_id,
                employees.c.employee_code,
                employees.c.full_name,
                detection_events.c.confidence,
                employees.c.deactivation_reason,
                employees.c.deactivated_at,
            )
            .select_from(
                detection_events.outerjoin(
                    cameras,
                    and_(
                        cameras.c.id == detection_events.c.camera_id,
                        cameras.c.tenant_id == detection_events.c.tenant_id,
                    ),
                ).outerjoin(
                    employees,
                    and_(
                        employees.c.id
                        == detection_events.c.former_match_employee_id,
                        employees.c.tenant_id == detection_events.c.tenant_id,
                    ),
                )
            )
            .where(
                detection_events.c.tenant_id == scope.tenant_id,
                detection_events.c.former_employee_match.is_(True),
                detection_events.c.captured_at >= start_utc,
                detection_events.c.captured_at < end_utc,
            )
            .order_by(detection_events.c.captured_at.desc())
        ).all()

    return [
        FormerEmployeeSighting(
            detection_event_id=int(r.id),
            captured_at=r.captured_at,
            camera_id=int(r.camera_id) if r.camera_id is not None else None,
            camera_name=str(r.camera_name) if r.camera_name is not None else None,
            former_employee_id=(
                int(r.former_match_employee_id)
                if r.former_match_employee_id is not None
                else None
            ),
            former_employee_code=(
                str(r.employee_code) if r.employee_code is not None else None
            ),
            former_employee_name=(
                str(r.full_name) if r.full_name is not None else None
            ),
            confidence=float(r.confidence) if r.confidence is not None else None,
            deactivation_reason=r.deactivation_reason,
            deactivated_at=r.deactivated_at,
        )
        for r in rows
    ]


@router.get("/former-employees-seen")
def former_employees_seen_report(
    user: Annotated[CurrentUser, ADMIN_OR_HR],
    from_date: Annotated[date, Query(alias="from")],
    to_date: Annotated[date, Query(alias="to")],
    format_: Annotated[Literal["json", "xlsx"], Query(alias="format")] = "json",
):
    """Return former-employee sightings between two dates (inclusive)."""

    if to_date < from_date:
        raise HTTPException(
            status_code=400, detail="`to` must be on or after `from`"
        )
    if (to_date - from_date).days > 366:
        raise HTTPException(
            status_code=400,
            detail="date range too large (max 366 days)",
        )

    scope = TenantScope(tenant_id=user.tenant_id)
    items = _query_sightings(scope, from_date=from_date, to_date=to_date)

    # Audit every export so an auditor can see who pulled the report.
    with get_engine().begin() as conn:
        write_audit(
            conn,
            tenant_id=scope.tenant_id,
            actor_user_id=user.id,
            action="report.former_employees_seen",
            entity_type="report",
            entity_id=None,
            after={
                "from": from_date.isoformat(),
                "to": to_date.isoformat(),
                "format": format_,
                "row_count": len(items),
            },
        )

    if format_ == "json":
        return FormerEmployeeSightingsOut(
            items=items,
            total=len(items),
            from_date=from_date,
            to_date=to_date,
        )

    # XLSX path.
    from openpyxl import Workbook  # noqa: PLC0415
    from openpyxl.styles import Font  # noqa: PLC0415

    wb = Workbook(write_only=False)
    ws = wb.active
    ws.title = "Former employees seen"
    bold = Font(bold=True)
    headers = [
        "Captured at (UTC)",
        "Camera",
        "Former employee code",
        "Former employee name",
        "Confidence",
        "Deactivation reason",
        "Deactivated at",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
    for i, row in enumerate(items, start=2):
        ws.cell(row=i, column=1, value=row.captured_at.isoformat())
        ws.cell(row=i, column=2, value=row.camera_name or "")
        ws.cell(row=i, column=3, value=row.former_employee_code or "")
        ws.cell(row=i, column=4, value=row.former_employee_name or "")
        ws.cell(
            row=i,
            column=5,
            value=row.confidence if row.confidence is not None else "",
        )
        ws.cell(row=i, column=6, value=row.deactivation_reason or "")
        ws.cell(
            row=i,
            column=7,
            value=row.deactivated_at.isoformat()
            if row.deactivated_at is not None
            else "",
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = (
        f"maugood-former-employees-{from_date.isoformat()}-to-{to_date.isoformat()}.xlsx"
    )
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )

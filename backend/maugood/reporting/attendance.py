"""Attendance Excel report builder.

Uses openpyxl in **write-only mode** so we don't materialise the entire
workbook in memory at once — for a multi-week, multi-employee report
that's the difference between a few MB transient memory and tens of MB.
The router streams the resulting bytes through ``StreamingResponse``.

Sheets are partitioned by ISO calendar week (e.g. ``2026-W17``) — the
pilot plan's choice; HR can fold a long range into manageable tabs
without re-running with narrower filters.

Columns: ``employee_code, name, date, in_time, out_time, total_hours,
late, early_out, short, overtime_minutes, policy``.
"""

from __future__ import annotations

from datetime import date, datetime, time
from io import BytesIO
from typing import Iterator, Optional

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from sqlalchemy import and_, select
from sqlalchemy.engine import Connection

from maugood.db import (
    attendance_records,
    departments,
    employees,
    shift_policies,
)
from maugood.tenants.scope import TenantScope

REPORT_COLUMNS: tuple[str, ...] = (
    "employee_code",
    "name",
    "date",
    "in_time",
    "out_time",
    "total_hours",
    "late",
    "early_out",
    "short",
    "overtime_minutes",
    "policy",
)


def _iso_week_label(d: date) -> str:
    iso = d.isocalendar()
    return f"{iso.year}-W{iso.week:02d}"


def _format_time(t: Optional[time]) -> str:
    return t.strftime("%H:%M:%S") if t is not None else ""


def _query_rows(
    conn: Connection,
    scope: TenantScope,
    *,
    start_date: date,
    end_date: date,
    department_ids: Optional[list[int]],
    employee_id: Optional[int],
) -> Iterator[tuple]:
    """Yield joined attendance rows ordered by ``(date asc, employee_code asc)``."""

    stmt = (
        select(
            employees.c.employee_code,
            employees.c.full_name,
            attendance_records.c.date,
            attendance_records.c.in_time,
            attendance_records.c.out_time,
            attendance_records.c.total_minutes,
            attendance_records.c.late,
            attendance_records.c.early_out,
            attendance_records.c.short_hours,
            attendance_records.c.absent,
            attendance_records.c.overtime_minutes,
            shift_policies.c.name.label("policy_name"),
        )
        .select_from(
            attendance_records.join(
                employees,
                and_(
                    employees.c.id == attendance_records.c.employee_id,
                    employees.c.tenant_id == attendance_records.c.tenant_id,
                ),
            )
            .join(
                departments,
                and_(
                    departments.c.id == employees.c.department_id,
                    departments.c.tenant_id == employees.c.tenant_id,
                ),
            )
            .join(
                shift_policies,
                and_(
                    shift_policies.c.id == attendance_records.c.policy_id,
                    shift_policies.c.tenant_id == attendance_records.c.tenant_id,
                ),
            )
        )
        .where(
            attendance_records.c.tenant_id == scope.tenant_id,
            attendance_records.c.date >= start_date,
            attendance_records.c.date <= end_date,
        )
        .order_by(
            attendance_records.c.date.asc(),
            employees.c.employee_code.asc(),
        )
    )
    if department_ids is not None:
        if not department_ids:
            return iter(())
        stmt = stmt.where(employees.c.department_id.in_(department_ids))
    if employee_id is not None:
        stmt = stmt.where(employees.c.id == employee_id)
    return conn.execute(stmt)


def build_xlsx(
    conn: Connection,
    scope: TenantScope,
    *,
    start_date: date,
    end_date: date,
    department_ids: Optional[list[int]] = None,
    employee_id: Optional[int] = None,
) -> tuple[bytes, int]:
    """Build the XLSX in-memory and return ``(bytes, row_count)``.

    Write-only mode keeps the workbook from holding every cell in
    memory; we still serialize to a single ``BytesIO`` for the
    response because openpyxl can't write directly to a streaming
    HTTP body. For pilot data volumes (a few weeks × ~100 employees)
    the buffer is well under a MB.
    """

    wb = Workbook(write_only=True)
    sheets: dict[str, object] = {}

    rows_written = 0
    for row in _query_rows(
        conn,
        scope,
        start_date=start_date,
        end_date=end_date,
        department_ids=department_ids,
        employee_id=employee_id,
    ):
        label = _iso_week_label(row.date)
        sheet = sheets.get(label)
        if sheet is None:
            sheet = wb.create_sheet(title=label)
            # Header row, bold-cell only on the first sheet would be
            # nice; write-only mode supports it via WriteOnlyCell. Keep
            # it simple — a plain header row is fine for pilot.
            sheet.append(list(REPORT_COLUMNS))
            sheets[label] = sheet
        total_hours = (
            round(row.total_minutes / 60.0, 2)
            if row.total_minutes is not None
            else ""
        )
        sheet.append(
            [
                row.employee_code,
                row.full_name,
                row.date.isoformat(),
                _format_time(row.in_time),
                _format_time(row.out_time),
                total_hours,
                bool(row.late),
                bool(row.early_out),
                bool(row.short_hours),
                int(row.overtime_minutes),
                row.policy_name,
            ]
        )
        rows_written += 1

    # Empty result still gets a single empty sheet so the operator
    # opens a file with the expected header row, not a corrupt
    # "no sheets" workbook.
    if not sheets:
        ws = wb.create_sheet(title="Attendance")
        ws.append(list(REPORT_COLUMNS))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), rows_written
